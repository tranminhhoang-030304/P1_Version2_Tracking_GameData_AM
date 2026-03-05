import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from api.etl_processor import run_etl_pipeline
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS

import psycopg2
from psycopg2.extras import RealDictCursor
import json
from datetime import datetime, timedelta, timezone
import time
import requests
import threading
import random
import os                       
from dotenv import load_dotenv
import re
import io
from collections import defaultdict

load_dotenv()

app = Flask(__name__)

# CORS Configuration: Cho phép FE gọi từ domain khác
# Development: cho phép tất cả
# Production: nên chỉ định domain cụ thể
allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")
if allowed_origins == ["*"]:
    # Development mode: cho phép tất cả
    CORS(app, resources={r"/*": {"origins": "*"}})
else:
    # Production mode: chỉ cho phép domain được chỉ định
    CORS(app, resources={
        r"/*": {
            "origins": allowed_origins,
            "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
            "allow_headers": ["Content-Type", "Authorization"],
            "credentials": True
        }
    })

# --- 3. SỬA CẤU HÌNH DATABASE (Lấy từ .env) ---
DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASS") 

# Kiểm tra an toàn: Nếu không đọc được pass thì báo lỗi
if not DB_PASS:
    print("⚠️  CẢNH BÁO: Chưa tìm thấy DB_PASS trong file .env")

# --- QUẢN LÝ TRẠNG THÁI ĐA LUỒNG (PARALLEL MODE) ---
# Thay vì khóa cả hệ thống, chỉ khóa từng App ID đang chạy
RUNNING_APPS = set()
APP_LOCK = threading.Lock()
JOB_STOP_EVENTS = {}

def is_app_busy(app_id):
    with APP_LOCK:
        return app_id in RUNNING_APPS
    
def try_lock_app(app_id):
    """
    Cố gắng khóa App để chạy Job. 
    Trả về True nếu khóa thành công (App đang rảnh).
    Trả về False nếu App đang bận chạy job khác.
    """
    with APP_LOCK:
        if app_id in RUNNING_APPS:
            return False
        RUNNING_APPS.add(app_id)
        return True
    
def unlock_app(app_id):
    """Giải phóng App sau khi chạy xong"""
    with APP_LOCK:
        RUNNING_APPS.discard(app_id)

def is_system_busy(): return False
def set_system_busy(busy, app_id=None, run_type=None): pass

def parse_date_param(date_str):
    # [FIX] Nếu FE gửi rác (undefined, null, rỗng), tự động trả về ngày hôm nay (Giờ VN)
    if not date_str or str(date_str).strip().lower() in ['undefined', 'null', 'none', '']:
        vn_tz = timezone(timedelta(hours=7))
        return datetime.now(vn_tz).strftime('%Y-%m-%d')
        
    try:
        # Thử format DD/MM/YYYY (Nếu FE gửi lên dạng này)
        return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
    except:
        # Nếu đã là YYYY-MM-DD thì giữ nguyên
        return str(date_str)

# --- [NEW] HÀM GIẢI PHÓNG DỮ LIỆU (FLATTEN & MERGE) ---
def strict_flatten_event(data):
    """
    1. Tìm các key chứa JSON string (như 'event_json', 'params').
    2. Parse chúng ra.
    3. Gộp (Merge) vào object cha.
    4. Xóa key gốc.
    5. Nếu lỗi -> Ném Exception để Fail Job.
    """
    # Danh sách các key thường chứa dữ liệu lồng nhau
    nested_keys_to_explode = ['event_json', 'params', 'data', 'attributes']
    
    # Copy để tránh ảnh hưởng biến gốc khi đang loop
    result = data.copy() if isinstance(data, dict) else {}
    
    for key in nested_keys_to_explode:
        if key in result:
            raw_value = result[key]
            
            # Chỉ xử lý nếu nó là String (cần parse) hoặc Dict (cần gộp)
            inner_data = None
            
            if isinstance(raw_value, str):
                # Thử Parse Strict
                if raw_value.strip().startswith('{'):
                    try:
                        inner_data = json.loads(raw_value)
                    except Exception as e:
                        raise ValueError(f"❌ STRICT MERGE FAIL: Key '{key}' contains invalid JSON. Error: {e}")
            elif isinstance(raw_value, dict):
                inner_data = raw_value
            
            # Nếu đã lấy được ruột, tiến hành Gộp và Xóa vỏ
            if isinstance(inner_data, dict):
                # 1. Gộp vào cha (Update) - Dữ liệu con sẽ đè lên cha nếu trùng tên (thường là mới hơn)
                result.update(inner_data)
                # 2. Xóa key lồng nhau đi (theo yêu cầu sếp)
                del result[key]
            
            # Nếu là string nhưng không phải JSON object (vd: "some text"), giữ nguyên hoặc xử lý tùy ý
            # Ở đây ta giữ nguyên nếu không parse được thành dict

    return result

# --- [UPDATED HELPER] GIẢI NÉN JSON TẦNG SÂU (SAFE MODE) ---
def recursive_json_unpack(data):
    """
    Hàm đệ quy thông minh:
    - Cố gắng biến String JSON -> Object/List.
    - NẾU LỖI: Trả về nguyên gốc (String) thay vì làm sập Job.
    """
    # 1. Nếu là String: Thử giải nén, nếu thất bại thì thôi
    if isinstance(data, str):
        stripped = data.strip()
        # Chỉ thử parse nếu trông nó giống JSON
        if (stripped.startswith('{') and stripped.endswith('}')) or \
           (stripped.startswith('[') and stripped.endswith(']')):
            try:
                parsed = json.loads(stripped)
                # Đệ quy tiếp vào bên trong (đề phòng lồng nhiều lớp)
                return recursive_json_unpack(parsed)
            except:
                # 🛑 QUAN TRỌNG: Nếu lỗi parse, TRẢ VỀ NGUYÊN GỐC (không raise lỗi)
                return data
        return data

    # 2. Nếu là Dict: Quét từng key
    if isinstance(data, dict):
        return {k: recursive_json_unpack(v) for k, v in data.items()}

    # 3. Nếu là List: Quét từng phần tử
    if isinstance(data, list):
        return [recursive_json_unpack(item) for item in data]

    # Các kiểu khác (int, float, None...) giữ nguyên
    return data

def universal_flatten(raw_input):
    """
    Hàm V95: Khoan sâu vào mọi ngóc ngách của JSON (Hỗ trợ n lớp lồng nhau).
    Trả về: Một dictionary phẳng chứa tất cả thông tin.
    """
    if not raw_input: return {}
    
    data = {}
    # Lớp 1: Parse từ DB (thường là string hoặc dict)
    try:
        if isinstance(raw_input, str):
            data = json.loads(raw_input)
        elif isinstance(raw_input, dict):
            data = raw_input.copy()
    except: return {}

    # Lớp 2: Kiểm tra các key chứa JSON lồng nhau thường gặp
    # Game 2 thường nhét dữ liệu vào key 'event_json' hoặc 'params'
    nested_keys = ['event_json', 'params', 'data', 'attributes']
    
    for key in nested_keys:
        if key in data and isinstance(data[key], str):
            try:
                inner = json.loads(data[key])
                if isinstance(inner, dict):
                    data.update(inner) # Gộp dữ liệu con ra ngoài
            except: pass
            
    # Lớp 3: Xử lý Double Encode (Trường hợp chuỗi bị mã hóa 2 lần)
    if isinstance(data, str): 
        try: data = json.loads(data)
        except: pass
        
    return data if isinstance(data, dict) else {}

def get_app_config(cur, app_id):
    """
    Hàm lấy cấu hình động từ Database.
    Nếu không có, trả về cấu hình mặc định (Fallback) để tránh lỗi.
    """
    try:
        cur.execute("SELECT config_json FROM analytics_config WHERE app_id = %s", (app_id,))
        row = cur.fetchone()
        if row and row['config_json']:
            return row['config_json']
    except Exception as e:
        print(f"Config Warning: {e}")
        # Quan trọng: Nếu query lỗi, phải rollback để không kẹt transaction sau này
        if cur.connection:
            cur.connection.rollback()

    # Cấu hình Mặc định (Fallback) nếu chưa setup DB
    return {
        "events": {
            "start": ["missionStart", "missionStart_Daily", "missionStart_WeeklyQuestTutor"],
            "win": ["missionComplete", "missionComplete_Daily", "missionComplete_WeeklyQuestTutor"],
            "progress": ["missionProgress"],
            "fail": ["missionFail", "missionFail_Daily", "missionFail_WeeklyQuestTutor"],
            "transaction": {
                "real_currency": ["iapSuccess", "firstIAP"], # <--- Đã thêm dấu phẩy
                "virtual_currency_exclude": ["iapSuccess", "firstIAP", "iapPurchase", "priceSpendLevel"], # <--- Đã thêm dấu phẩy
                "offer_and_reward": ["FirstReward", "adsRewardComplete", "iapOfferGet", "dailyReward"]
            }
        },
        "boosters": [ # <--- Sửa ngoặc nhọn { thành ngoặc vuông [
            {"key": "booster_Hammer", "name": "Hammer 🔨", "type": "booster"},
            {"key": "booster_Magnet", "name": "Magnet 🧲", "type": "booster"},
            {"key": "booster_Add", "name": "Add Moves ➕", "type": "booster"},
            {"key": "booster_Unlock", "name": "Unlock 🔓", "type": "booster"},
            {"key": "booster_Clear", "name": "Clear 🧹", "type": "booster"},
            {"key": "revive_boosterClear", "name": "Revive 💖", "type": "revive"}
        ], # <--- Sửa ngoặc nhọn } thành ngoặc vuông [
        "currency": {
            "real": ["VND", "USD", "₫", "$"], # <--- Đã thêm dấu phẩy
            "virtual": ["Coin"]
        }
    }

def smart_parse_json(raw_input):
    """
    Hàm thông minh để xử lý trường hợp JSON bị lồng 2 lớp.
    Ví dụ: "{\"event_json\": \"{\\\"levelID\\\": 1...}\"}"
    """
    if not raw_input: 
        return {}
    
    try:
        # Lớp 1: Nếu là string thì parse ra dict, nếu là dict rồi thì giữ nguyên
        parsed_data = json.loads(raw_input) if isinstance(raw_input, str) else raw_input
        
        # Lớp 2: Kiểm tra xem bên trong có key 'event_json' chứa string JSON nữa không (Lỗi double encode)
        if isinstance(parsed_data, dict) and 'event_json' in parsed_data:
            inner_value = parsed_data['event_json']
            if isinstance(inner_value, str):
                try:
                    inner_json = json.loads(inner_value)
                    # Gộp dữ liệu bên trong ra ngoài (Flatten)
                    parsed_data.update(inner_json)
                except:
                    pass # Nếu không parse được lớp trong thì thôi
                    
        return parsed_data
    except Exception:
        return {}

def get_db():
    try:
        conn = psycopg2.connect(
            host=os.getenv('DB_HOST'),
            database=os.getenv('DB_NAME'),
            user=os.getenv('DB_USER'),
            password=os.getenv('DB_PASS') or os.getenv('DB_PASSWORD'),
            port=os.getenv('DB_PORT')
        )
        return conn
    except Exception as e:
        print(f"❌ DB Connection Error: {e}")
        return None

# ==========================================
# PHẦN 1: CORE FUNCTIONS (TẠO JOB & CẬP NHẬT)
# ==========================================

def create_etl_job(app_id, run_type, scheduled_time, date_since, date_until):
    conn = get_db()
    if not conn: return
    cur = conn.cursor()
    try:
        # Check trùng
        cur.execute("""
            SELECT id FROM job_history 
            WHERE app_id = %s 
              AND date_since = %s 
              AND run_type = %s
              AND status IN ('pending', 'Processing', 'Running')
        """, (app_id, date_since, run_type))
        
        if cur.fetchone(): return 

        cur.execute("""
            INSERT INTO job_history (app_id, run_type, scheduled_time, status, created_at, date_since, date_until)
            VALUES (%s, %s, %s, 'pending', NOW(), %s, %s) RETURNING id
        """, (app_id, run_type, scheduled_time, date_since, date_until))
        
        conn.commit()
        print(f"🎫 {run_type.capitalize()}: Đã tạo vé Job cho App {app_id} (Window: {date_since} -> {date_until})")
    except Exception as e:
        print(f"❌ Create Job Error: {e}")
    finally:
        cur.close()
        conn.close()

def update_job_status(job_id, status, message=None, inc_retry=False):
    conn = get_db()
    if not conn: return
    cur = conn.cursor()
    try:
        sql = "UPDATE etl_jobs SET status = %s, updated_at = NOW(), message = %s"
        if inc_retry: sql += ", retry_count = retry_count + 1"
        sql += " WHERE id = %s"
        cur.execute(sql, (status, message, job_id))
        conn.commit()
    finally:
        cur.close()
        conn.close()

# ==========================================
# PHẦN 2: WORKER THÔNG MINH (ĐÃ SỬA LỖI TIME & INSERT TRƯỚC)
# ==========================================
# --- HÀM PHỤ TRỢ: GHI LOG VÀO DB ---
def append_log_to_db(hist_id, new_log_line):
    """Nối thêm log vào dòng lịch sử đang chạy"""   
    if not hist_id: return
    try:
        conn = get_db()
        cur = conn.cursor()
        # Dùng toán tử || để nối chuỗi trong PostgreSQL
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_entry = f"\n[{timestamp}] {new_log_line}"
        
        cur.execute("""
            UPDATE job_history 
            SET logs = COALESCE(logs, '') || %s, updated_at = NOW()
            WHERE id = %s
        """, (log_entry, hist_id))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ Error appending log: {e}")

def transform_events_to_level_analytics(app_id, events):
    """
    [UPDATED] Transform missionStart / missionComplete / missionFail
    Sử dụng smart_parse_json để xử lý lỗi lồng JSON.
    """
    if not events:
        return

    conn = get_db()
    cur = conn.cursor()

    # 1. Gom event theo (user_id, level_id)
    sessions = {}

    for e in events:
        try:
            event_name = e.get("event_name")
            raw_json = smart_parse_json(e.get("event_json"))
            level_id = (raw_json.get("levelID") or 
                        raw_json.get("missionID") or 
                        raw_json.get("level_display") or 
                        raw_json.get("level_display_origin"))
            user_id = (raw_json.get("userID") or "Guest" or
                       raw_json.get("uuid"))
            # Nếu vẫn không lấy được level_id, bỏ qua event này
            if not level_id: continue
            session_key = f"{user_id}_{level_id}"
            # Timestamp xử lý an toàn
            try:
                ts_val = int(e.get("event_timestamp"))
                ts = datetime.fromtimestamp(ts_val)
            except:
                ts = datetime.now()

            if session_key not in sessions:
                sessions[session_key] = {
                    "app_id": app_id,
                    "user_id": user_id,
                    "level_id": level_id,
                    "start_time": None,
                    "end_time": None,
                    "status": "DROP",
                    "total_cost": 0
                }

            s = sessions[session_key]

            # Logic tính toán giữ nguyên, chỉ đảm bảo raw_json đã sạch
            if event_name == "missionStart":
                s["start_time"] = ts

            elif event_name == "missionComplete":
                s["end_time"] = ts
                s["status"] = "WIN"
                for k, v in raw_json.items():
                    if k.startswith("booster_") or k.startswith("revive_"):
                        try: s["total_cost"] += int(v)
                        except: pass

            elif event_name == "missionFail":
                s["end_time"] = ts
                s["status"] = "FAIL"
                for k, v in raw_json.items():
                    if k.startswith("booster_") or k.startswith("revive_"):
                        try: s["total_cost"] += int(v)
                        except: pass

        except Exception as ex:
            print(f"Transform error skipping row: {ex}")

    # 2. Insert vào level_analytics
    for s in sessions.values():
        start_time = s["start_time"]
        end_time = s["end_time"]
        duration = 0
        if start_time and end_time:
            duration = int((end_time - start_time).total_seconds())

        # Chỉ insert nếu có dữ liệu hợp lệ (Tránh rác)
        try:
            cur.execute("""
                INSERT INTO level_analytics
                (app_id, session_id, user_id, level_name, status, duration, start_time, total_cost, created_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            """, (
                s["app_id"],
                f"{s['user_id']}_{s['level_id']}",
                s["user_id"],
                f"Level {s['level_id']}", # Format tên Level đẹp hơn
                s["status"],
                duration,
                start_time,
                s["total_cost"]
            ))
        except Exception as insert_err:
            print(f"Insert Analytics Error: {insert_err}")

    conn.commit()
    conn.close()

def execute_job_logic(job_id, app_id, retry_count, run_type, retry_job_id):
    """
    Worker V121: Perfect Log (Giao diện cũ + Logic mới)
    """
    hist_id = None
    try:
        conn = get_db()
        if not conn: return
        cur = conn.cursor()
        
        # 1. Tạo Log History
        cur.execute("INSERT INTO job_history (app_id, start_time, status, run_type, logs, total_events) VALUES (%s, NOW(), 'Running', 'schedule', '', 0) RETURNING id", (app_id,))
        hist_id = cur.fetchone()[0]
        conn.commit()
        
        def log(msg):
            ts = datetime.now().strftime("[%H:%M:%S]")
            print(f"{ts} [App {app_id}] {msg}")
            append_log_to_db(hist_id, msg)

        def cancel_job():
            log("🛑 Đã nhận tín hiệu STOP. Đang dọn dẹp và hủy Job an toàn...")
            update_job_status(job_id, 'cancelled') # Đổi etl_jobs thành cancelled
            try:
                c_fin = get_db(); cr_fin = c_fin.cursor()
                cr_fin.execute("UPDATE job_history SET end_time=NOW(), status='Cancelled' WHERE id=%s", (hist_id,))
                c_fin.commit(); c_fin.close()
            except: pass

        # [LOG START] Hiển thị rõ số lần Retry hệ thống
        log(f"▶️ Start Job #{job_id} (System Retry: {retry_count})...")

        # 2. Lấy thông tin Job & App
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM etl_jobs WHERE id = %s", (job_id,))
        job = cur.fetchone()
        conn.close() 

        conn_thread = get_db()
        cur_thread = conn_thread.cursor(cursor_factory=RealDictCursor)
        cur_thread.execute("SELECT * FROM apps WHERE id = %s", (app_id,))
        app_info = cur_thread.fetchone()
        conn_thread.close() 
        
        if not app_info:
            log("❌ App missing."); update_job_status(job_id, 'failed'); return

        clean_app_id = str(app_info['app_id']).strip()
        clean_token = str(app_info['api_token']).strip()
        
        # 🟢 3. KIỂM TRA THỜI GIAN & LOG WINDOW (Logic V109 + Log chi tiết)
        try:
            u_start = datetime.strptime(str(job['date_since']), '%Y-%m-%d %H:%M:%S')
            u_end = datetime.strptime(str(job['date_until']), '%Y-%m-%d %H:%M:%S')
            
            # Check tương lai (Giữ logic V109 cho an toàn)
            now_utc = datetime.utcnow()
            if u_end > now_utc:
                log(f"⛔ TỪ CHỐI: Job yêu cầu quét tương lai (End: {u_end} > Now: {now_utc}).")
                update_job_status(job_id, 'failed', "Error: Future Time")
                conn_fail = get_db(); cur_fail = conn_fail.cursor()
                cur_fail.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                conn_fail.commit(); conn_fail.close()
                return

            # [LOG WINDOW] Hiển thị cả giờ VN và UTC ngay đầu Job (Chỉ 1 lần)
            vn_start = u_start + timedelta(hours=7)
            vn_end = u_end + timedelta(hours=7)
            log(f"🕒 Scanning Window: VN[{vn_start.strftime('%H:%M')} - {vn_end.strftime('%H:%M')}] (UTC: {u_start.strftime('%H:%M')} - {u_end.strftime('%H:%M')})")

        except Exception as e:
            log(f"⚠️ Lỗi parse thời gian: {e}")

        try:
            req_start = (datetime.strptime(str(job['date_since']), '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
            req_end = (datetime.strptime(str(job['date_until']), '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
        except:
            req_start = str(job['date_since'])
            req_end = str(job['date_until'])

        # --- CHUẨN BỊ GỌI API ---
        clean_app_id = str(app_info['app_id']).strip()
        clean_token = str(app_info['api_token']).strip()
        
        wanted_fields = [
            "event_datetime", "event_json", "event_name", "event_receive_datetime", "event_receive_timestamp", "event_timestamp", "session_id", "installation_id", 
            "appmetrica_device_id", "city", "connection_type", "country_iso_code", "device_ipv6", "device_locale", "device_manufacturer", "device_model", 
            "device_type", "google_aid", "ios_ifa", "ios_ifv", "mcc", "mnc", "operator_name", "original_device_model", "os_name", "os_version", "profile_id",
            "windows_aid", "app_build_number", "app_package_name", "app_version_name", "application_id"
        ]

        # --- SETUP REQUEST ---
        url = "https://api.appmetrica.yandex.com/logs/v1/export/events.json"
        params = {
            "application_id": clean_app_id,
            "date_since": req_start,
            "date_until": req_end,
            "fields": ",".join(wanted_fields)
        }
        headers = {"Authorization": f"OAuth {clean_token}"}
        
        # --- LOGIC RETRY ---
        max_attempts = 18 
        poll_interval = 180 # 3p
        
        for attempt in range(1, max_attempts + 1):
            if JOB_STOP_EVENTS.get(hist_id) and JOB_STOP_EVENTS[hist_id].is_set():
                cancel_job(); break
            
            # [LOG ATTEMPT] Hiển thị rõ lần thử thứ mấy (1/18)
            log(f"📡 Requesting AppMetrica ({attempt}/{max_attempts})...")
            
            try:
                response = requests.get(url, params=params, headers=headers, stream=True, timeout=600)
                
                # [TRƯỜNG HỢP 1]: THÀNH CÔNG CÓ DATA (200 OK)
                if response.status_code == 200:
                    data = response.json()
                    events = data.get('data', [])
                    count = len(events)
                    
                    log(f"✅ Downloaded {count} events. Saving...")
                    log(f"🕵️‍♂️ [DEBUG] AppMetrica gửi về: {count} events (Raw).")
                    
                    conn_ins = get_db(); cur_ins = conn_ins.cursor()
                    vals = []
                    for e in events:
                        evt_json = json.dumps(e)
                        try: ts = datetime.fromtimestamp(int(e.get('event_timestamp')))
                        except: ts = datetime.now()
                        vals.append((app_id, e.get('event_name'), evt_json, 1, ts, hist_id))
                    
                    if vals:
                        cur_ins.executemany("""
                        INSERT INTO event_logs (app_id, event_name, event_json, count, created_at, job_id) 
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, vals)
                    conn_ins.commit(); conn_ins.close()
                    
                    log("🔄 Processing Analytics...")
                    try: transform_events_to_level_analytics(app_id, events)
                    except Exception as te: log(f"⚠️ Transform: {te}")

                    log(f"🎉 SUCCESS. Imported {count} events.")
                    update_job_status(job_id, 'completed', f"OK. {count} events.")
                    
                    conn_fin = get_db(); cur_fin = conn_fin.cursor()
                    cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Success', total_events=%s WHERE id=%s", (count, hist_id))
                    conn_fin.commit(); conn_fin.close()
                    break 
                
                # [TRƯỜNG HỢP 2]: ĐANG GOM DATA (202 ACCEPTED)
                elif response.status_code == 202:
                    log(f"  ⏳ HTTP 202: Data preparing... Waiting {poll_interval}s...")
                    try:
                        c_re = get_db(); cr_re = c_re.cursor()
                        cr_re.execute("UPDATE job_history SET retry_count = %s WHERE id = %s", (attempt, hist_id))
                        c_re.commit(); c_re.close()
                    except: pass

                    stop_event = JOB_STOP_EVENTS.get(hist_id)
                    if stop_event:
                        if stop_event.wait(poll_interval): cancel_job(); break
                    else:
                        time.sleep(poll_interval) # Backup: Nếu mất event chặn thì vẫn bắt buộc phải ngủ
                    continue 

                # [TRƯỜNG HỢP 3]: KẸT XE CHỜ 3 PHÚT (429 HOẶC ENQUEUED)
                elif response.status_code == 429 or "enqueued" in response.text.lower():
                    log(f"  ❌ Lỗi 429: Hàng đợi AppMetrica đang đầy. Đánh dấu FAILED, vui lòng dùng nút Retry sau!")
                    
                    # Cập nhật bảng cũ (etl_jobs)
                    update_job_status(job_id, 'failed', f"Rate Limit 429")
                    
                    # Cập nhật DB thành Failed và kết thúc thời gian chạy
                    try:
                        conn_fin = get_db(); cur_fin = conn_fin.cursor()
                        cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                        conn_fin.commit(); conn_fin.close()
                    except: pass
                    return
                
                # [TRƯỜNG HỢP 4]: LỖI KHÁC (Cái đuôi bị mất của bạn) - Ép chết Job nếu lỗi API sai cú pháp
                else:
                    err_text = response.text.strip()[:200]
                    log(f"❌ FATAL ERROR {response.status_code}: {err_text}")
                    update_job_status(job_id, 'failed', f"API Error {response.status_code}")
                    
                    conn_fin = get_db(); cur_fin = conn_fin.cursor()
                    cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                    conn_fin.commit(); conn_fin.close()
                    return # Rút ống thở luôn, không vòng lặp nữa

            # [NGOẠI LỆ]: ĐỨT MẠNG, TIMEOUT...
            except Exception as e_req:
                log(f"⚠️ Request Error: {e_req}")
                stop_event = JOB_STOP_EVENTS.get(hist_id)
                if stop_event:
                    if stop_event.wait(60): cancel_job(); break
                else:
                    time.sleep(60)
        
        else:
            if JOB_STOP_EVENTS.get(hist_id) and not JOB_STOP_EVENTS[hist_id].is_set():
                log("❌ Job Failed (Max Retries).")
                update_job_status(job_id, 'failed', "Max Retries")
                # ... (update failed db)
                conn_fin = get_db(); cur_fin = conn_fin.cursor()
                cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                conn_fin.commit(); conn_fin.close()

    except Exception as e:
        print(f"Critical: {e}")
        if hist_id: append_log_to_db(hist_id, f"❌ Crash: {e}")
    finally:
        print(f"🔓 App {app_id} Free.")
        unlock_app(app_id)

def execute_manual_job_logic(hist_id, app_id, retry_count=0):
    try:
        conn = get_db()
        if not conn: return
        
        # 1. Update trạng thái thành Running ngay lập tức
        cur = conn.cursor()
        cur.execute("UPDATE job_history SET status='Running', start_time=NOW() WHERE id=%s", (hist_id,))
        conn.commit()
        
        def log(msg):
            ts = datetime.now().strftime("[%H:%M:%S]")
            print(f"{ts} [App {app_id} - MANUAL] {msg}")
            append_log_to_db(hist_id, msg)

        def cancel_job():
            log("🛑 Đã nhận tín hiệu STOP Manual Job.")
            try:
                c_fin = get_db(); cr_fin = c_fin.cursor()
                cr_fin.execute("UPDATE job_history SET end_time=NOW(), status='Cancelled' WHERE id=%s", (hist_id,))
                c_fin.commit(); c_fin.close()
            except: pass

        log(f"▶️ Start MANUAL Job (History ID: {hist_id})...")

        # 2. Lấy thông tin App và Ngày tháng trực tiếp từ bảng job_history
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM apps WHERE id = %s", (app_id,))
        app_info = cur.fetchone()
        
        cur.execute("SELECT date_since, date_until FROM job_history WHERE id = %s", (hist_id,))
        job_time = cur.fetchone()
        conn.close() 

        if not app_info or not job_time:
            log("❌ App hoặc Job Time missing."); return

        # Gán biến giả lập giống hệt etl_jobs để phần code phía dưới chạy bình thường
        job = {
            'date_since': job_time['date_since'],
            'date_until': job_time['date_until']
        }
        
        # 🟢 3. KIỂM TRA THỜI GIAN & LOG WINDOW (Logic V109 + Log chi tiết)
        try:
            u_start = datetime.strptime(str(job['date_since']), '%Y-%m-%d %H:%M:%S')
            u_end = datetime.strptime(str(job['date_until']), '%Y-%m-%d %H:%M:%S')
            
            # Check tương lai (Giữ logic V109 cho an toàn)
            now_utc = datetime.utcnow()
            if u_end > now_utc:
                log(f"⛔ TỪ CHỐI: Job yêu cầu quét tương lai (End: {u_end} > Now: {now_utc}).")
                #update_job_status(job_id, 'failed', "Error: Future Time")
                conn_fail = get_db(); cur_fail = conn_fail.cursor()
                cur_fail.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                conn_fail.commit(); conn_fail.close()
                return

            # [LOG WINDOW] Hiển thị cả giờ VN và UTC ngay đầu Job (Chỉ 1 lần)
            vn_start = u_start + timedelta(hours=7)
            vn_end = u_end + timedelta(hours=7)
            log(f"🕒 Scanning Window: VN[{vn_start.strftime('%H:%M')} - {vn_end.strftime('%H:%M')}] (UTC: {u_start.strftime('%H:%M')} - {u_end.strftime('%H:%M')})")

        except Exception as e:
            log(f"⚠️ Lỗi parse thời gian: {e}")

        try:
            req_start = (datetime.strptime(str(job['date_since']), '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
            req_end = (datetime.strptime(str(job['date_until']), '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
        except:
            req_start = str(job['date_since'])
            req_end = str(job['date_until'])

        # --- CHUẨN BỊ GỌI API ---
        clean_app_id = str(app_info['app_id']).strip()
        clean_token = str(app_info['api_token']).strip()
        
        wanted_fields = [
            "event_datetime", "event_json", "event_name", "event_receive_datetime", "event_receive_timestamp", "event_timestamp", "session_id", "installation_id", 
            "appmetrica_device_id", "city", "connection_type", "country_iso_code", "device_ipv6", "device_locale", "device_manufacturer", "device_model", 
            "device_type", "google_aid", "ios_ifa", "ios_ifv", "mcc", "mnc", "operator_name", "original_device_model", "os_name", "os_version", "profile_id",
            "windows_aid", "app_build_number", "app_package_name", "app_version_name", "application_id"
        ]

        # --- SETUP REQUEST ---
        url = "https://api.appmetrica.yandex.com/logs/v1/export/events.json"
        params = {
            "application_id": clean_app_id,
            "date_since": req_start,
            "date_until": req_end,
            "fields": ",".join(wanted_fields)
        }
        headers = {"Authorization": f"OAuth {clean_token}"}
        
        # --- LOGIC RETRY ---
        max_attempts = 18 
        poll_interval = 180 # 3p
        
        for attempt in range(1, max_attempts + 1):
            if JOB_STOP_EVENTS.get(hist_id) and JOB_STOP_EVENTS[hist_id].is_set():
                cancel_job(); break
            
            # [LOG ATTEMPT] Hiển thị rõ lần thử thứ mấy (1/18)
            log(f"📡 Requesting AppMetrica ({attempt}/{max_attempts})...")
            
            try:
                response = requests.get(url, params=params, headers=headers, stream=True, timeout=600)
                
                # [TRƯỜNG HỢP 1]: THÀNH CÔNG CÓ DATA (200 OK)
                if response.status_code == 200:
                    data = response.json()
                    events = data.get('data', [])
                    count = len(events)
                    
                    log(f"✅ Downloaded {count} events. Saving...")
                    log(f"🕵️‍♂️ [DEBUG] AppMetrica gửi về: {count} events (Raw).")
                    
                    conn_ins = get_db(); cur_ins = conn_ins.cursor()
                    vals = []
                    for e in events:
                        evt_json = json.dumps(e)
                        try: ts = datetime.fromtimestamp(int(e.get('event_timestamp')))
                        except: ts = datetime.now()
                        vals.append((app_id, e.get('event_name'), evt_json, 1, ts, hist_id))
                    
                    if vals:
                        cur_ins.executemany("""
                        INSERT INTO event_logs (app_id, event_name, event_json, count, created_at, job_id) 
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, vals)
                    conn_ins.commit(); conn_ins.close()
                    
                    log("🔄 Processing Analytics...")
                    try: transform_events_to_level_analytics(app_id, events)
                    except Exception as te: log(f"⚠️ Transform: {te}")

                    log(f"🎉 SUCCESS. Imported {count} events.")
                    #update_job_status(job_id, 'completed', f"OK. {count} events.")
                    
                    conn_fin = get_db(); cur_fin = conn_fin.cursor()
                    cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Success', total_events=%s WHERE id=%s", (count, hist_id))
                    conn_fin.commit(); conn_fin.close()
                    break 
                
                # [TRƯỜNG HỢP 2]: ĐANG GOM DATA (202 ACCEPTED)
                elif response.status_code == 202:
                    log(f"  ⏳ HTTP 202: Data preparing... Waiting {poll_interval}s...")
                    try:
                        c_re = get_db(); cr_re = c_re.cursor()
                        cr_re.execute("UPDATE job_history SET retry_count = %s WHERE id = %s", (attempt, hist_id))
                        c_re.commit(); c_re.close()
                    except: pass

                    stop_event = JOB_STOP_EVENTS.get(hist_id)
                    if stop_event:
                        if stop_event.wait(poll_interval): cancel_job(); break
                    else:
                        time.sleep(poll_interval) # Backup: Nếu mất event chặn thì vẫn bắt buộc phải ngủ
                    continue 

                # [TRƯỜNG HỢP 3]: KẸT XE CHỜ 3 PHÚT (429 HOẶC ENQUEUED)
                elif response.status_code == 429 or "enqueued" in response.text.lower():
                    log(f"  ❌ Lỗi 429: Hàng đợi AppMetrica đang đầy. Đánh dấu FAILED, vui lòng dùng nút Retry sau!")
                    
                    # Cập nhật DB thành Failed và kết thúc thời gian chạy (Manual chỉ dùng hist_id)
                    try:
                        conn_fin = get_db(); cur_fin = conn_fin.cursor()
                        cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                        conn_fin.commit(); conn_fin.close()
                    except: pass
                    return 
                
                # [TRƯỜNG HỢP 4]: LỖI KHÁC (Cái đuôi bị mất của bạn) - Ép chết Job nếu lỗi API sai cú pháp
                else:
                    err_text = response.text.strip()[:200]
                    log(f"❌ FATAL ERROR {response.status_code}: {err_text}")
                    #update_job_status(job_id, 'failed', f"API Error {response.status_code}")
                    
                    conn_fin = get_db(); cur_fin = conn_fin.cursor()
                    cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                    conn_fin.commit(); conn_fin.close()
                    return # Rút ống thở luôn, không vòng lặp nữa

            # [NGOẠI LỆ]: ĐỨT MẠNG, TIMEOUT...
            except Exception as e_req:
                log(f"⚠️ Request Error: {e_req}")
                stop_event = JOB_STOP_EVENTS.get(hist_id)
                if stop_event:
                    if stop_event.wait(60): cancel_job(); break
                else:
                    time.sleep(60)
        
        else:
            if JOB_STOP_EVENTS.get(hist_id) and not JOB_STOP_EVENTS[hist_id].is_set():
                log("❌ Job Failed (Max Retries).")
                #update_job_status(job_id, 'failed', "Max Retries")
                # ... (update failed db)
                conn_fin = get_db(); cur_fin = conn_fin.cursor()
                cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                conn_fin.commit(); conn_fin.close()

    except Exception as e:
        print(f"Critical: {e}")
        if hist_id: append_log_to_db(hist_id, f"❌ Crash: {e}")
    finally:
        print(f"🔓 App {app_id} Free.")
        unlock_app(app_id)

def worker_process_jobs():
    # Kiểm tra xem có Worker nào đang chạy không
    if is_system_busy(): return

    conn = get_db()
    if not conn: return
    
    # -------------------------------------------------------------
    # [NEW] LOGIC CHỌN JOB THÔNG MINH (HỖ TRỢ HẸN GIỜ)
    # Thay vì chỉ lấy 'pending', ta lấy thêm điều kiện scheduled_at
    # -------------------------------------------------------------
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Ưu tiên 1: Kiểm tra bảng chung job_history (Nơi chứa Manual Job & Scheduled Job)
    # Lấy job pending MÀ (không hẹn giờ HOẶC đã đến giờ hẹn)
    cur.execute("""
        SELECT * FROM job_history 
        WHERE status = 'pending' 
          AND (scheduled_at IS NULL OR scheduled_at <= (NOW() AT TIME ZONE 'UTC'))
        ORDER BY created_at ASC LIMIT 1
    """)
    job_history_row = cur.fetchone()
    
    # Ưu tiên 2: Nếu job_history rỗng, kiểm tra bảng cũ etl_jobs (Legacy Auto Job)
    # Để đảm bảo không mất chức năng cũ của Auto Scheduler
    job_legacy = None
    if not job_history_row:
        cur.execute("""
            SELECT * FROM etl_jobs 
            WHERE status IN ('pending', 'processing') 
            ORDER BY created_at ASC LIMIT 1
        """)
        job_legacy = cur.fetchone()
    
    cur.close() 
    conn.close()

    # --- HỢP NHẤT DỮ LIỆU JOB (ADAPTER) ---
    # Mục tiêu: Dù lấy từ bảng nào thì cũng map về một cấu trúc 'job' chuẩn
    job = None
    source_table = 'none'

    if job_history_row:
        job = job_history_row
        source_table = 'history'
    elif job_legacy:
        job = job_legacy
        source_table = 'legacy'
    else:
        return # Không có việc gì làm

    job_id = job['id']
    app_id = job['app_id']

    # --- [RETRY LOGIC V2 - FIX DATA RANGE] ---
    # Mục tiêu: Job cũ quét từ A đến B -> Job Retry cũng phải quét từ A đến B (y hệt)
    if job.get('run_type') == 'retry' and job.get('retry_job_id'):
        print(f"🕵️‍♂️ Retry Detect: Đang khôi phục cấu hình từ Job #{job['retry_job_id']}...")
        try:
            # 1. Kết nối DB để lấy thông tin gốc
            conn_fix = get_db()
            cur_fix = conn_fix.cursor(cursor_factory=RealDictCursor)
            
            # 2. Lấy chính xác khoảng thời gian (Window) của Job cũ
            # date_since / date_until là khoảng dữ liệu (VD: 16:58 - 18:03)
            cur_fix.execute("""
                SELECT date_since, date_until, app_id 
                FROM job_history 
                WHERE id = %s
            """, (job['retry_job_id'],))
            
            old_job_data = cur_fix.fetchone()
            cur_fix.close(); conn_fix.close()

            # 3. Ốp dữ liệu cũ vào Job hiện tại
            if old_job_data:
                # [QUAN TRỌNG] Ép kiểu về string để tránh lỗi datetime object
                if old_job_data.get('date_since'):
                    job['date_since'] = str(old_job_data['date_since'])
                
                if old_job_data.get('date_until'):
                    job['date_until'] = str(old_job_data['date_until'])
                
                # Đảm bảo app_id khớp nhau (phòng hờ)
                if old_job_data.get('app_id'):
                    job['app_id'] = old_job_data['app_id']

                print(f"✅ RETRY FIXED: Đã khôi phục Data Range chuẩn: {job['date_since']} -> {job['date_until']}")
            else:
                print(f"⚠️ RETRY WARNING: Không tìm thấy Job cũ #{job['retry_job_id']} trong DB. Sẽ chạy theo tham số hiện tại.")

        except Exception as e:
            print(f"❌ RETRY ERROR: Lỗi khi khôi phục Job cũ: {e}")
            # Nếu lỗi, code sẽ chạy tiếp với tham số mặc định

    # [NEW] CẬP NHẬT TRẠNG THÁI (PROCESSING)
    # Nếu là job từ bảng history (Manual/Scheduled), ta update ngay trạng thái để tránh worker khác lấy trùng
    if source_table == 'history':
        try:
            conn_upd = get_db(); cur_upd = conn_upd.cursor()
            cur_upd.execute("UPDATE job_history SET status='Processing', start_time=NOW() WHERE id=%s", (job_id,))
            conn_upd.commit(); conn_upd.close()
        except: pass

    # 2. KHỞI TẠO HISTORY (LOGIC CŨ)
    # Nếu job lấy từ 'history' (Manual), thì hist_id chính là nó luôn.
    # Nếu job lấy từ 'etl_jobs' (Legacy), thì phải tạo dòng mới trong job_history.
    hist_id = None
    try:
        if source_table == 'history':
            hist_id = job_id # Dùng luôn dòng hiện tại
        else:
            # Logic cũ cho etl_jobs
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT id FROM job_history WHERE app_id = %s AND status IN ('Running', 'Processing') ORDER BY start_time DESC LIMIT 1", (app_id,))
            row = cur.fetchone()
            if row:
                hist_id = row[0]
            else:
                cur.execute("""
                    INSERT INTO job_history (app_id, start_time, status, run_type, logs, total_events, date_since, date_until)
                    VALUES (%s, NOW(), 'Processing', 'schedule', '', 0, %s, %s)
                    RETURNING id
                """, (app_id, job['date_since'], job['date_until']))
                hist_id = cur.fetchone()[0]
            conn.commit(); conn.close()
    except Exception as e:
        print(f"❌ Init Error: {e}")
        set_system_busy(False)
        return

    stop_event = threading.Event()
    if hist_id:
        JOB_STOP_EVENTS[hist_id] = stop_event

    # Hàm log cục bộ (Giữ nguyên)
    def log(msg):
        print(msg)
        if hist_id: append_log_to_db(hist_id, msg)

    try:
        # 3. FIX TIMEZONE & LẤY APP INFO (GIỮ NGUYÊN 100%)
        try:
            # Timezone Logic (Hiển thị +7)
            utc_s = datetime.strptime(str(job['date_since']), '%Y-%m-%d %H:%M:%S')
            utc_e = datetime.strptime(str(job['date_until']), '%Y-%m-%d %H:%M:%S')
            vn_start = utc_s + timedelta(hours=7)
            vn_end = utc_e + timedelta(hours=7)
            log(f" 📅 Date: {vn_start.strftime('%Y-%m-%d')} | 🕒 Window: VN[{vn_start.strftime('%H:%M')} - {vn_end.strftime('%H:%M')}]")
        except: pass

        # Lấy Info App
        conn = get_db(); cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM apps WHERE id = %s", (app_id,))
        app_info = cur.fetchone()
        cur.close(); conn.close()

        if not app_info:
            if source_table == 'legacy': update_job_status(job_id, 'failed', 'App deleted')
            return

        # 4. CHUẨN BỊ GỌI API (GIỮ NGUYÊN 100%)
        clean_app_id = str(app_info['app_id']).strip()
        clean_token = str(app_info['api_token']).strip()
        
        wanted_fields = [
            "event_datetime", "event_json", "event_name", "event_receive_datetime", "event_receive_timestamp", "event_timestamp", "session_id", "installation_id", 
            "appmetrica_device_id", "city", "connection_type", "country_iso_code", "device_ipv6", "device_locale", "device_manufacturer", "device_model", 
            "device_type", "google_aid", "ios_ifa", "ios_ifv", "mcc", "mnc", "operator_name", "original_device_model", "os_name", "os_version", "profile_id",
            "windows_aid", "app_build_number", "app_package_name", "app_version_name", "application_id"
        ]
        
        url = "https://api.appmetrica.yandex.com/logs/v1/export/events.json"
        params = {
            "application_id": clean_app_id,
            "date_since": str(job['date_since']), 
            "date_until": str(job['date_until']),
            "fields": ",".join(wanted_fields),
        }
        headers = {"Authorization": f"OAuth {clean_token}"}

        db_retry_count = job.get('retry_count') or 0
        # --- LOGIC RETRY (GIỮ NGUYÊN 100%) ---
        max_polling_attempts = 18   
        polling_interval = 180      
        
        for attempt in range(1, max_polling_attempts + 1):
            total_attempts = db_retry_count + attempt
            log(f"  📡 Kết nối AppMetrica (Attempt {attempt}/{max_polling_attempts})...")
            
            try:
                response = requests.get(url, params=params, headers=headers, stream=True, timeout=600)
                
                # TRƯỜNG HỢP 1: CÓ HÀNG (200 OK)
                if response.status_code == 200:
                    log("  ✅ Data Ready (200 OK). Downloading & Processing...")
                    
                    data = response.json()
                    events = data.get('data', [])
                    event_count = len(events)
                    
                    # --- BATCH INSERT & UPSERT (ĐÃ FIX THEO YÊU CẦU CỦA SẾP) ---
                    conn_ins = get_db(); cur_ins = conn_ins.cursor()
                    vals = []
                    
                    for event in events:
                        evt_name = event.get('event_name', 'unknown')
                        final_json_str = "{}"
                        
                        # 1. Xử lý JSON (Giữ nguyên)
                        try:
                            clean_s1 = recursive_json_unpack(event) 
                            final_flat = strict_flatten_event(clean_s1) 
                            final_json_str = json.dumps(final_flat, ensure_ascii=False)
                        except:
                            try: final_json_str = json.dumps(event, ensure_ascii=False)
                            except: final_json_str = "{}"

                        # 2. Xử lý UUID & Timestamp Key (String)
                        # [NÂNG CẤP]: Nếu uuid rỗng, thử lấy installation_id lấp vào
                        uuid_val = str(event.get('uuid') or '')
                        if not uuid_val or uuid_val == 'None':
                             uuid_val = str(event.get('installation_id') or '')
                             
                        raw_ts_val = str(event.get('event_timestamp') or '')

                        # 3. Xử lý Thời gian (FINAL: UTC CHUẨN - KHÔNG CỘNG TRỪ)
                        try: 
                            raw_ts = event.get('event_timestamp', 0)
                            ts_val = float(raw_ts)
                            
                            # Milliseconds detection
                            if ts_val > 99999999999: ts_val = ts_val / 1000.0
                            
                            # [CHỐT HẠ] Lấy giờ UTC gốc và lưu thẳng vào DB.
                            # Ví dụ: Sự kiện lúc 00:00 VN -> UTC là 17:00.
                            # Lưu 17:00 vào DB.
                            # Frontend lấy 17:00 + 7 = 00:00 (Chuẩn đét!)
                            
                            ts = datetime.fromtimestamp(ts_val, timezone.utc).replace(tzinfo=None)
                            
                            # TUYỆT ĐỐI KHÔNG dùng dòng này: ts = ts - timedelta(hours=7)
                            # TUYỆT ĐỐI KHÔNG dùng dòng này: ts = ts + timedelta(hours=7)
                            
                        except: 
                            ts = datetime.utcnow()

                        # 4. Append
                        vals.append((app_id, evt_name, final_json_str, 1, ts, hist_id, uuid_val, raw_ts_val))
                    
                    if vals:
                        # [CHIẾN THUẬT AUDIT TOÀN DIỆN]
                        # 1. Key duy nhất vẫn là 4 trường cũ (để phát hiện trùng).
                        # 2. Khi trùng:
                        #    - Cất Job ID cũ vào job_id_old.
                        #    - Cất JSON cũ vào event_json_old.
                        #    - Ghi Job ID mới và JSON mới vào cột chính.
                        
                        query = """
                            INSERT INTO event_logs (app_id, event_name, event_json, count, created_at, job_id, uuid, raw_timestamp) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            ON CONFLICT (app_id, event_name, raw_timestamp, uuid) 
                            DO UPDATE SET 
                                -- [BACKUP DỮ LIỆU CŨ]
                                event_json_old = event_logs.event_json,   -- Lưu vết nội dung
                                job_id_old = event_logs.job_id,           -- Lưu vết Job cũ (Theo ý sếp)
                                
                                -- [GHI DỮ LIỆU MỚI]
                                event_json = EXCLUDED.event_json,
                                job_id = EXCLUDED.job_id,
                                created_at = EXCLUDED.created_at,
                                count = event_logs.count
                        """
                        cur_ins.executemany(query, vals)
                        conn_ins.commit(); conn_ins.close()
                        
                        print(f"  💾 Saved {len(vals)} events (Upsert Mode with Full Audit).")
                    
                    # Transform (GIỮ NGUYÊN)
                    try: transform_events_to_level_analytics(app_id, events)
                    except: pass

                    # Success Finish
                    conn = get_db(); cur = conn.cursor()
                    cur.execute("UPDATE job_history SET end_time=NOW(), status='Success', total_events=%s, success_count=%s WHERE id=%s", (event_count, event_count, hist_id))
                    conn.commit(); conn.close()
                    
                    if source_table == 'legacy':
                        update_job_status(job_id, 'completed', f"Done. {event_count} events.")
                        
                    log(f"  🎉 Job Completed. Imported: {event_count} events.")
                    return 
                
                # TRƯỜNG HỢP 2: CHƯA CÓ HÀNG (202 ACCEPTED)
                elif response.status_code == 202:
                    log(f"  ⏳ HTTP 202: Data preparing... Waiting {polling_interval}s...")
                    
                    if source_table == 'legacy':
                        update_job_status(job_id, 'processing', f"Waiting Data (Try {attempt}/{max_polling_attempts})...")
                    
                    # [UPDATED] Update cả retry_count cho bảng history
                    try:
                        c_re = get_db(); cr_re = c_re.cursor()
                        cr_re.execute("UPDATE job_history SET retry_count = %s WHERE id = %s", (attempt, hist_id))
                        c_re.commit(); c_re.close()
                    except: pass

                    if stop_event.wait(polling_interval):
                        log("🛑 Đã nhận tín hiệu STOP! Rút ống thở Worker ngay lập tức!")
                        break # Phá vỡ vòng lặp 18 lần
                        
                    continue 

                elif response.status_code == 429 or "enqueued" in response.text.lower():
                    log(f"  ❌ Lỗi 429: Hàng đợi AppMetrica đang đầy. Đánh dấu FAILED, vui lòng dùng nút Retry sau!")
                    
                    if source_table == 'legacy':
                        update_job_status(job_id, 'failed', f"Rate Limit 429")
                    
                    # Cập nhật DB thành Failed và kết thúc thời gian chạy
                    try:
                        conn_fin = get_db(); cur_fin = conn_fin.cursor()
                        cur_fin.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                        conn_fin.commit(); conn_fin.close()
                    except: pass

                    return

                    if stop_event.wait(180):
                        log("🛑 Đã nhận tín hiệu STOP!")
                        break
                    continue

                # TRƯỜNG HỢP 3: LỖI
                else:
                    err_text = response.text.strip()[:200]
                    log(f"❌ FATAL ERROR {response.status_code}: {err_text}")
                    if source_table == 'legacy':
                        update_job_status(job_id, 'failed', f"API Error {response.status_code}")
                    
                    conn = get_db(); cur = conn.cursor()
                    cur.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                    conn.commit(); conn.close()
                    return 

            except Exception as e_req:
                log(f"⚠️ Request Error: {e_req}")
                time.sleep(60)
        
        # TIMEOUT SAU 18 LẦN
        log("❌ TIMEOUT: AppMetrica did not return data after max retries.")
        if source_table == 'legacy':
            update_job_status(job_id, 'failed', 'Timeout (202 Loop)')
        
        conn = get_db(); cur = conn.cursor()
        cur.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
        conn.commit(); conn.close()

    except Exception as e:
        log(f"❌ Worker Exception: {str(e)}")
        if source_table == 'legacy':
            update_job_status(job_id, 'failed', str(e))
        
        conn = get_db(); cur = conn.cursor()
        if hist_id: cur.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
        conn.commit(); conn.close()
    
    finally:
        set_system_busy(False)
        if hist_id and hist_id in JOB_STOP_EVENTS:
            del JOB_STOP_EVENTS[hist_id]

def run_worker_loop(worker_id):
    # Lệch nhịp khởi động: Bác sĩ 1 vào làm ngay, Bác sĩ 2 uống cà phê 5s rồi mới vào
    # Tránh việc 2 bác sĩ cùng lao vào giật 1 tờ phiếu khám trong Database
    time.sleep(worker_id * 5) 
    print(f"🚀 Worker Thread [{worker_id}] Started...")
    
    while True:
        try:
            worker_process_jobs()
        except Exception as e:
            print(f"❌ Worker [{worker_id}] Loop Error: {e}")
        
        # Làm xong 1 việc thì nghỉ 30s rồi tìm việc tiếp
        time.sleep(30)

def run_scheduler_loop():
    print("🚀 Auto Scheduler V2.1 (Timezone Fixed) Started...")
    
    # [VACCINE SERVER]: Ép định nghĩa múi giờ VN (+7) để bất chấp server đặt ở quốc gia nào
    vn_tz = timezone(timedelta(hours=7))

    while True:
        try:
            # Lấy giờ VN hiện tại, bỏ đi tzinfo để tính toán cộng trừ cho dễ
            now = datetime.now(vn_tz).replace(tzinfo=None)
            
            conn = get_db()
            if not conn:
                time.sleep(60); continue
                
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT * FROM apps WHERE is_active = true") 
            apps = cur.fetchall()

            for app in apps:
                app_id = app['id']
                
                # --- CẤU HÌNH ---
                interval = app.get('interval_minutes', 60) or 60
                sch_str = app.get('schedule_time', '00:00')
                
                # --- TÍNH TOÁN GIỜ CHẠY ---
                try: h, m = map(int, sch_str.split(':'))
                except: h=0; m=0
                
                anchor = now.replace(hour=h, minute=m, second=0, microsecond=0)
                if anchor > now: anchor -= timedelta(days=1)
                
                delta_seconds = (now - anchor).total_seconds()
                cycles_passed = int(delta_seconds // (interval * 60))
                
                # Giờ dự kiến chạy của chu kỳ này
                expected_run_time = anchor + timedelta(minutes=cycles_passed * interval)
                
                if expected_run_time <= now <= (expected_run_time + timedelta(minutes=2)):
                    
                    # --- [CẤU HÌNH GỐI ĐẦU - OVERLAP STRATEGY] ---
                    delay_minutes = 90
                    cycle_minutes = interval 
                    overlap_minutes = 5      
                    
                    # --- [PHẪU THUẬT CHỐNG JOB SINH ĐÔI] ---
                    # Dùng chính mốc expected_run_time (Cố định) thay vì now (Trôi) 
                    # Đảm bảo dù thức dậy bao nhiêu lần trong 2 phút cũng chỉ ra 1 khung giờ DUY NHẤT!
                    end_dt_vn = expected_run_time - timedelta(minutes=delay_minutes)
                    start_dt_vn = end_dt_vn - timedelta(minutes=cycle_minutes + overlap_minutes)
                    
                    # Trả lại sự trong sáng (Lưu UTC vào DB)
                    end_dt_utc = end_dt_vn - timedelta(hours=7)
                    start_dt_utc = start_dt_vn - timedelta(hours=7)
                    
                    s_str = start_dt_utc.strftime('%Y-%m-%d %H:%M:%S')
                    e_str = end_dt_utc.strftime('%Y-%m-%d %H:%M:%S')

                    # --- [FIX] KIỂM TRA TRÙNG LẶP DỰA TRÊN DATA RANGE ---
                    check_cur = conn.cursor()
                    check_cur.execute("""
                        SELECT id FROM job_history 
                        WHERE app_id = %s 
                          AND date_since = %s 
                          AND date_until = %s
                    """, (app_id, s_str, e_str))
                    existing = check_cur.fetchone()
                    check_cur.close()
                    
                    if existing:
                        # Đã có job cho khung giờ này rồi -> Bỏ qua, không spam log
                        pass 
                    else:
                        print(f"⏰ [App {app_id}] Trigger Auto Job for range (VN Time): {s_str} -> {e_str}")
                        create_etl_job(
                            app_id=app['id'],
                            run_type='schedule',
                            scheduled_time=expected_run_time,
                            date_since=s_str,  
                            date_until=e_str  
                        )
            
            cur.close(); conn.close()
        except Exception as e:
            print(f"❌ Scheduler Error: {e}")
     
        time.sleep(90)

# PHẦN 4: LOGIC CHẠY TAY (MANUAL) - [FIXED RETRY LOGIC]
def perform_manual_etl(app_id, run_type='manual', is_demo=False, retry_job_id=None):
    if is_system_busy():
        print(f"❌ System BUSY. Skip run for App {app_id}.")
        return

    set_system_busy(True, app_id, run_type)
    hist_id = None

    try:
        conn = get_db()
        if not conn: return
        
        # 1. TẠO RECORD HISTORY
        msg_start = f"🚀 Starting {run_type.upper()} run..."
        if retry_job_id: msg_start += f" (Retry of Job #{retry_job_id})"
        
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO job_history (app_id, start_time, status, run_type, logs, total_events) 
            VALUES (%s, NOW(), 'Running', %s, %s, 0) 
            RETURNING id
        """, (app_id, run_type, f"[{datetime.now().strftime('%H:%M:%S')}] {msg_start}"))
        hist_id = cur.fetchone()[0]
        conn.commit()
        
        stop_event = threading.Event()
        JOB_STOP_EVENTS[hist_id] = stop_event

        def log(msg):
            print(msg)
            append_log_to_db(hist_id, msg)

        # 2. LẤY CẤU HÌNH APP
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM apps WHERE id=%s", (app_id,))
        app = cur.fetchone()
        
        if not app:
            log("❌ Error: App ID not found.")
            cur.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
            conn.commit()
            return 

        # 3. CẤU HÌNH THỜI GIAN
        date_since = None
        date_until = None
        
        # --- CASE 1: RETRY CHUẨN (LOGIC V100 - ĐỌC TRỰC TIẾP TỪ DB) ---
        # "Bác sĩ" kê đơn: Bỏ Regex, đọc thẳng cột date_since/date_until trong bảng job_history
        if run_type == 'retry' and retry_job_id:
            try:
                # 1. Truy vấn hồ sơ gốc của Job cũ
                cur.execute("SELECT date_since, date_until, app_id FROM job_history WHERE id = %s", (retry_job_id,))
                old_row = cur.fetchone()
                
                if old_row and old_row.get('date_since') and old_row.get('date_until'):
                    # 2. Lấy dữ liệu chuẩn (Ép kiểu string để code bên dưới chạy không lỗi)
                    date_since = str(old_row['date_since'])
                    date_until = str(old_row['date_until'])
                    
                    # [Quan trọng] Cập nhật App ID nếu cần (để đảm bảo tính toàn vẹn)
                    if old_row.get('app_id'):
                        app_id = old_row['app_id']

                    log(f"✅ RETRY FIXED: Khôi phục chính xác Data Range từ DB: {date_since} -> {date_until}")
                
                else:
                    # 3. Trường hợp Job quá cũ (trước khi sửa code) -> DB chưa có date_since
                    log(f"❌ RETRY FAILED: Job #{retry_job_id} không có thông tin date_since trong DB (Job cũ). Vui lòng chạy Manual.")
                    # Đánh dấu Job hiện tại là Failed luôn
                    cur.execute("UPDATE job_history SET end_time=NOW(), status='Failed', logs=CONCAT(logs, '\n❌ Retry Failed: No metadata in DB') WHERE id=%s", (hist_id,))
                    conn.commit()
                    return

            except Exception as e:
                log(f"❌ RETRY EXCEPTION: Lỗi khi đọc DB: {e}")
                cur.execute("UPDATE job_history SET end_time=NOW(), status='Failed' WHERE id=%s", (hist_id,))
                conn.commit()
                return

        # --- CASE 2: DEMO / MANUAL ---
        if not date_since:
            now = datetime.now()
            delay_minutes = 45 if run_type == 'demo' else 90
            duration_minutes = 15 if run_type == 'demo' else 60
            
            end_dt = now - timedelta(minutes=delay_minutes)
            start_dt = end_dt - timedelta(minutes=duration_minutes)
            
            date_since = start_dt.strftime('%Y-%m-%d %H:%M:%S')
            date_until = end_dt.strftime('%Y-%m-%d %H:%M:%S')
            log(f"⚙️ MANUAL MODE: {date_since} -> {date_until}")

        log(f" 🕒 Scanning Window: {date_since} -> {date_until}")
        
        # 4. GỌI API APPMETRICA (Dùng .strip() an toàn)
        clean_app_id = str(app['app_id']).strip()
        clean_token = str(app['api_token']).strip()

        try:
            req_start = (datetime.strptime(date_since, '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
            req_end = (datetime.strptime(date_until, '%Y-%m-%d %H:%M:%S') + timedelta(hours=7)).strftime('%Y-%m-%d %H:%M:%S')
        except:
            req_start = date_since
            req_end = date_until

        url = "https://api.appmetrica.yandex.com/logs/v1/export/events.json"
        params = {
            "application_id": clean_app_id,
            "date_since": req_start,
            "date_until": req_end,
            "fields": "event_name,event_timestamp,event_json",
        }
        headers = {"Authorization": f"OAuth {clean_token}"}
        
        status = "Failed"; total_events = 0
        
        for i in range(18): # 18 retries
            if stop_event.is_set():
                log("🛑 USER STOPPED PROCESS.")
                status = "Cancelled"; break

            log(f"📡 Requesting AppMetrica (Attempt {i+1}/18)...")
            resp = requests.get(url, params=params, headers=headers)
            
            if resp.status_code == 200:
                data = resp.json().get('data', [])
                total_events = len(data)
                log(f"✅ Success! Received {total_events} events. Importing...")
                
                conn_insert = get_db(); cur_insert = conn_insert.cursor()
                values = []
                for d in data:
                    try: ts = datetime.fromtimestamp(int(d.get('event_timestamp')))
                    except: ts = datetime.now()
                    values.append((app_id, d.get('event_name', 'unknown'), json.dumps(d), 1, ts))
                
                cur_insert.executemany("INSERT INTO event_logs (app_id, event_name, event_json, count, created_at) VALUES (%s,%s,%s,%s,%s)", values)
                conn_insert.commit(); conn_insert.close()
                
                status = "Success"; log(f"🎉 Done. Imported {total_events} events."); break
            
            elif resp.status_code == 202:
                log(f"⏳ Server 202. Waiting 180s..."); 
                if stop_event.wait(180): status = "Cancelled"; break
            else:
                log(f"❌ Error {resp.status_code}"); status = "Failed"; break
        
        # Cập nhật kết quả cuối cùng
        conn_end = get_db(); cur_end = conn_end.cursor()
        cur_end.execute("UPDATE job_history SET end_time=NOW(), status=%s, total_events=%s WHERE id=%s", (status, total_events, hist_id))
        conn_end.commit(); conn_end.close()
    
    except Exception as e:
        log(f"❌ Critical Error: {str(e)}")
    finally:
        set_system_busy(False)

@app.route("/monitor/history", methods=['GET'])
def get_history():
    app_id = request.args.get('app_id') 
    # [MỚI] Lấy tham số phân trang an toàn
    try:
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 30)) 
    except:
        page = 1; limit = 30
        
    offset = (page - 1) * limit
    conn = get_db()
    
    # Return cấu trúc chuẩn ngay cả khi không có DB
    if not conn: 
        return jsonify({
            "data": [], 
            "pagination": {"current_page": page, "total_pages": 0, "total_records": 0}
        })

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # 1. Xây dựng mệnh đề WHERE
        where_clause = ""
        params_count = []
        if app_id: 
            where_clause = "WHERE h.app_id = %s"
            params_count.append(app_id)

        # 2. Đếm tổng số records
        cur.execute(f"SELECT COUNT(*) as total FROM job_history h {where_clause}", tuple(params_count))
        total_records = cur.fetchone()['total']
        total_pages = (total_records + limit - 1) // limit

        # 3. Lấy dữ liệu 
        query = f"""
            SELECT 
                h.id, h.app_id, h.status, h.run_type, h.logs, 
                COALESCE(h.total_events, 0) as total_events,
                to_char(h.start_time, 'YYYY-MM-DD HH24:MI:SS') as start_time,
                to_char(h.end_time, 'YYYY-MM-DD HH24:MI:SS') as end_time,
                to_char(h.scheduled_at, 'YYYY-MM-DD HH24:MI:SS') as scheduled_at,
                to_char(h.date_since + interval '7 hours', 'YYYY-MM-DD HH24:MI:SS') as date_since,
                to_char(h.date_until + interval '7 hours', 'YYYY-MM-DD HH24:MI:SS') as date_until,
                
                a.name as app_name 
            FROM job_history h 
            JOIN apps a ON h.app_id = a.id 
            {where_clause}
            ORDER BY h.created_at DESC 
            LIMIT %s OFFSET %s
        """
        
        params_data = params_count + [limit, offset]
        cur.execute(query, tuple(params_data))
        res = cur.fetchall()

        # --- [FIX LỖI DURATION & DATA CHECK] ---
        for row in res:
            # Tính duration an toàn
            duration_str = "-"
            try:
                if row['start_time'] and row['end_time']:
                    s = datetime.strptime(row['start_time'], '%Y-%m-%d %H:%M:%S')
                    e = datetime.strptime(row['end_time'], '%Y-%m-%d %H:%M:%S')
                    diff = e - s
                    total_seconds = int(diff.total_seconds())
                    minutes = total_seconds // 60
                    seconds = total_seconds % 60
                    duration_str = f"{minutes}m {seconds}s"
                elif row['status'] == 'Running':
                     duration_str = "Running..."
            except: pass
            
            row['duration'] = duration_str

        return jsonify({
            "data": res,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "total_records": total_records
            }
        })

    except Exception as e:
        print(f"❌ Monitor History Error: {e}")
        # [QUAN TRỌNG] Trả về object pagination đầy đủ (số 0) thay vì rỗng để FE không sập
        return jsonify({
            "data": [], 
            "pagination": {
                "current_page": page, 
                "total_pages": 0, 
                "total_records": 0
            }
        })    
    finally: conn.close()

@app.route("/monitor/purge", methods=['DELETE'])
def purge_history():
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM job_history")
        conn.commit()
        return jsonify({"msg": "History Cleared"})
    finally: conn.close()

@app.route("/apps", methods=['GET', 'POST'])
def handle_apps():
    conn = get_db()
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        if request.method == 'GET':
            cur.execute("SELECT * FROM apps ORDER BY id ASC")
            return jsonify(cur.fetchall())
        else:
            d = request.json
            cur.execute("INSERT INTO apps (name, app_id, api_token, is_active, schedule_time, interval_minutes) VALUES (%s, %s, %s, %s, %s, %s)", 
                        (d['name'], d['app_id'], d['api_token'], d['is_active'], d.get('schedule_time', '12:00'), d.get('interval_minutes', 60)))
            conn.commit()
            return jsonify({"msg": "Created"})
    finally: conn.close()

@app.route("/apps/<int:id>", methods=['PUT', 'DELETE'])
def update_app(id):
    conn = get_db()
    try:
        cur = conn.cursor()
        if request.method == 'PUT':
            d = request.json
            cur.execute("UPDATE apps SET name=%s, app_id=%s, api_token=%s, is_active=%s, schedule_time=%s, interval_minutes=%s WHERE id=%s", 
                        (d['name'], d['app_id'], d['api_token'], d['is_active'], d.get('schedule_time', '12:00'), d.get('interval_minutes', 60), id))
            conn.commit()
            return jsonify({"msg": "Updated"})
        elif request.method == 'DELETE':
            cur.execute("DELETE FROM event_logs WHERE app_id=%s", (id,))
            cur.execute("DELETE FROM job_history WHERE app_id=%s", (id,))
            cur.execute("DELETE FROM etl_jobs WHERE app_id=%s", (id,))
            cur.execute("DELETE FROM apps WHERE id=%s", (id,))
            conn.commit()
            return jsonify({"msg": "Deleted"})
    finally: conn.close()

@app.route("/etl/run/<int:app_id>", methods=['POST'])
def run_etl_api(app_id):
    # --- CODE MỚI ---
    data = request.json
    run_type = data.get('run_type', 'manual') # Lấy loại chạy (manual/retry/demo)
    retry_job_id = data.get('retry_job_id')   # Lấy ID của job cũ nếu là retry
    
    is_demo = (run_type == 'demo')
    
    if is_system_busy():
         return jsonify({"status": "error", "message": "System is busy processing another job. Please skip this cycle."}), 409

    # Truyền thêm retry_job_id vào hàm xử lý
    threading.Thread(target=perform_manual_etl, args=(app_id, run_type, is_demo, retry_job_id)).start()
    return jsonify({"status": "started", "mode": run_type})

@app.route("/dashboard/<int:app_id>", methods=['GET'])
def get_dashboard(app_id):
    conn = get_db()
    if not conn: return jsonify({"success": False}), 500
    
    start_date = parse_date_param(request.args.get('start_date'))
    end_date = parse_date_param(request.args.get('end_date'))

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. CẤU HÌNH DISPLAY NAME & ICON
        # Map từ tên key (lower) sang tên hiển thị đẹp
        DISPLAY_MAP = {
            # Game 1 (Woolen Yarn)
            "add": "Add Moves ➕",
            "addmoves": "Add Moves ➕",
            "hammer": "Hammer 🔨",
            "magnet": "Magnet 🧲",
            "unlock": "Unlock 🔓",
            "clear": "Clear 🧹",
            "boosterclear": "Revive 💖", # Đổi tên theo yêu cầu
            "revive": "Revive 💖",
            
            # Game 2 (Loopy Bus)
            "bubble": "Bubble 🫧",
            "shuffle": "Shuffle 🔀",
            "ufo": "UFO 🛸"
        }

        # 2. CẤU HÌNH GIÁ (GIỮ NGUYÊN LOGIC CŨ)
        BOOSTER_PRICES = {}
        if app_id == 2: 
            BOOSTER_PRICES = {"bubble": 100, "shuffle": 80, "ufo": 150}
        else:
            BOOSTER_PRICES = {
                "hammer": 120, "magnet": 80, "add": 60, "addmoves": 60, 
                "unlock": 190, "clear": 120, "boosterclear": 120, "revive": 190
            }

        try:
            cfg = get_app_config(cur, app_id)
            if cfg and 'boosters' in cfg:
                for b in cfg['boosters']:
                    if isinstance(b, dict):
                        k = b.get('key','').replace('booster_','').replace('revive_','').lower()
                        if app_id != 2 and k in ['shuffle', 'undo', 'bubble', 'ufo']: continue
                        if app_id == 2 and k not in ['bubble', 'shuffle', 'ufo']: continue
                        BOOSTER_PRICES[k] = int(b.get('price', 100))
        except: pass

        # 3. QUERY & CALCULATE
        start_events = set(["missionStart", "missionStart_Daily", "level_start", "level_loading_start"])
        fail_events = set(["missionFail", "missionFail_Daily", "level_fail", "level_lose"])
        iap_events = set(["iapSuccess", "firstIAP", "iapPurchase", "purchase_verified", "iapOfferGet", "IAPOFFERCLOSE"]) 

        where = "WHERE app_id = %s"; params = [app_id]
        if start_date: where += " AND created_at >= %s"; params.append(start_date + " 00:00:00")
        if end_date: where += " AND created_at <= %s"; params.append(end_date + " 23:59:59")

        cur.execute(f"SELECT event_name, event_json FROM event_logs {where}", tuple(params))
        rows = cur.fetchall()

        import re
        def clean_money(val):
            if not val: return 0.0
            s_val = str(val)
            if s_val.count('.') > 1: return 0.0
            try:
                s = re.sub(r'[^\d.,-]', '', s_val)
                return float(s) if s else 0.0
            except: return 0.0

        overview = { "real_revenue": 0.0, "virtual_sink": 0, "total_plays": 0, "fail_count": 0, "total_time": 0.0 }
        event_dist = {}
        booster_map = {}
        level_stats = {} 
        
        for r in rows:
            evt = r['event_name']
            data = universal_flatten(r['event_json'])
            
            event_dist[evt] = event_dist.get(evt, 0) + 1
            if evt in start_events: overview['total_plays'] += 1
            if evt in fail_events: overview['fail_count'] += 1

            t_play = data.get('timeplay') or data.get('timePlay') or data.get('duration')
            if t_play:
                try: overview['total_time'] += float(t_play)
                except: pass

            if evt in iap_events:
                val = data.get('price') or data.get('revenue') or data.get('amount')
                overview['real_revenue'] += clean_money(val)
            
            coin_added = 0
            raw_coin = (data.get('coin_spent') or data.get('cost') or data.get('priceSpendLevel') or data.get('coinCost'))
            if not raw_coin and evt.lower() == 'pricespendlevel':
                 for v in data.values():
                     v_c = clean_money(v)
                     if v_c > 0 and v_c < 100000: raw_coin = v_c; break
            if raw_coin: coin_added = int(clean_money(raw_coin))
            
            if coin_added == 0:
                for k, v in data.items():
                    if ('booster' in k or 'revive' in k) and str(v).isdigit() and int(v) > 0:
                        clean_key = k.replace('booster_', '').replace('revive_', '').lower()
                        if clean_key in BOOSTER_PRICES:
                            coin_added += (int(v) * BOOSTER_PRICES[clean_key])

            overview['virtual_sink'] += coin_added

            for k, v in data.items():
                if ('booster' in k or 'revive' in k) and str(v).isdigit() and int(v) > 0:
                    clean = k.replace('booster_', '').replace('revive_', '').lower()
                    if clean in BOOSTER_PRICES:
                        booster_map[clean] = booster_map.get(clean, 0) + int(v)

            candidates = []
            for k in ['levelID', 'level_display', 'missionID', 'dayChallenge']:
                val = data.get(k)
                if val and str(val).isdigit(): candidates.append(int(val))
            if not candidates:
                match = re.search(r'(?:levelID|level_display|missionID)[^0-9]{1,10}(\d+)', r['event_json'])
                if match: candidates.append(int(match.group(1)))
            
            if candidates:
                lvl_num = max(candidates)
                if 0 < lvl_num < 5000:
                    lvl_key = str(lvl_num)
                    if lvl_key not in level_stats:
                        level_stats[lvl_key] = { "start": 0, "fail": 0, "revenue": 0.0 }
                    
                    ls = level_stats[lvl_key]
                    if evt in start_events: ls['start'] += 1
                    if evt in fail_events: ls['fail'] += 1
                    if evt in iap_events:
                        ls['revenue'] += clean_money(data.get('price') or data.get('revenue'))
                    if coin_added > 0:
                        ls['revenue'] += (coin_added / 1000.0)

        # 4. OUTPUT
        balance_chart = []
        for lvl, s in level_stats.items():
            fr = round((s['fail'] / s['start']) * 100, 1) if s['start'] > 0 else 0
            if s['start'] > 0: 
                balance_chart.append({
                    "name": f"Lv.{lvl}",
                    "level_index": int(lvl),
                    "fail_rate": fr,
                    "revenue": round(s['revenue'], 2),
                    "sessions": s['start']
                })
        balance_chart.sort(key=lambda x: x['level_index'])

        chart_data = [{"name": k, "value": v} for k, v in event_dist.items()]
        chart_data.sort(key=lambda x: x['value'], reverse=True)
        chart_data = chart_data[:20]

        # [FIX V130] GOM NHÓM THEO TÊN HIỂN THỊ (MERGE)
        merged_b = {}
        for k, v in booster_map.items():
            display_name = DISPLAY_MAP.get(k, k.capitalize())
            merged_b[display_name] = merged_b.get(display_name, 0) + v
        
        booster_stats = [{"name": k, "value": v} for k, v in merged_b.items()]
        booster_stats.sort(key=lambda x: x['value'], reverse=True)

        fail_rate_global = round((overview['fail_count'] / overview['total_plays']) * 100, 1) if overview['total_plays'] > 0 else 0.0
        avg_time_sec = round(overview['total_time'] / overview['total_plays'], 1) if overview['total_plays'] > 0 else 0

        return jsonify({
            "success": True,
            "overview": {
                "cards": {
                    "revenue": round(overview['real_revenue'], 2),
                    "active_users": overview['total_plays'],
                    "avg_fail_rate": fail_rate_global,
                    "total_spent": overview['virtual_sink'],
                    "avg_time": avg_time_sec
                },
                "chart_main": chart_data,
                "booster_chart": booster_stats,
                "balance_chart": balance_chart
            }
        })

    except Exception as e:
        print(f"Error Dashboard V129: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally: conn.close()

@app.route("/api/levels/<int:app_id>", methods=['GET'])
def get_levels(app_id):
    conn = get_db()
    if not conn: return jsonify([])
    try:
        cur = conn.cursor()
        # Chỉ lấy cột JSON để tối ưu
        cur.execute("SELECT event_json FROM event_logs WHERE app_id = %s", (app_id,))
        rows = cur.fetchall()
        
        levels = set()
        import re

        for r in rows:
            json_str = r[0]
            data = universal_flatten(json_str)
            
            # [LOGIC V116: TÌM ỨNG VIÊN LỚN NHẤT]
            candidates = []
            
            # 1. Quét qua các key tiềm năng
            keys_to_check = ['levelID', 'level_display', 'missionID', 'dayChallenge']
            for k in keys_to_check:
                val = data.get(k)
                if val and str(val).isdigit():
                    candidates.append(int(val))
            
            # 2. Regex Fallback
            if not candidates:
                match = re.search(r'(?:levelID|level_display|missionID)[^0-9]{1,10}(\d+)', json_str)
                if match:
                    candidates.append(int(match.group(1)))
            
            # 3. Chọn Level chuẩn
            if candidates:
                max_lvl = max(candidates)
                if 0 <= max_lvl <= 5000: # Giới hạn 5000 để lọc rác
                    levels.add(max_lvl)
        
        # Sort số -> string
        sorted_levels = sorted(list(levels))
        return jsonify([str(l) for l in sorted_levels])

    except Exception as e:
        print(f"Error get_levels: {e}")
        return jsonify([])
    finally: conn.close()


@app.route("/dashboard/<int:app_id>/level-detail", methods=['GET'])
def get_level_detail(app_id):
    safe_response = {
        "success": True, 
        "metrics": {"total_plays":0, "win_rate":0, "arpu":0, "avg_balance":0, "top_item":"None"},
        "funnel": [], "booster_usage": [], "cost_distribution": [],
        "logs": {"data": [], "pagination": {"current": 1, "total_pages": 0, "total_records": 0}}
    }

    try:
        level_id = request.args.get('level_id')
        start_date = parse_date_param(request.args.get('start_date'))
        end_date = parse_date_param(request.args.get('end_date'))
        try: page = int(request.args.get('page', 1)); limit = int(request.args.get('limit', 50))
        except: page=1; limit=50
        offset = (page - 1) * limit

        conn = get_db()
        if not conn: return jsonify(safe_response), 500
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. DISPLAY MAP & CONFIG
        DISPLAY_MAP = {
            "add": "Add Moves ➕", "addmoves": "Add Moves ➕",
            "hammer": "Hammer 🔨", "magnet": "Magnet 🧲",
            "unlock": "Unlock 🔓", "clear": "Clear 🧹",
            "boosterclear": "Revive 💖", "revive": "Revive 💖",
            "bubble": "Bubble 🫧", "shuffle": "Shuffle 🔀", "ufo": "UFO 🛸"
        }

        BOOSTER_PRICES = {}
        if app_id == 2:
            BOOSTER_PRICES = {"bubble": 100, "shuffle": 80, "ufo": 150}
        else:
            BOOSTER_PRICES = {
                "hammer": 120, "magnet": 80, "add": 60, "addmoves": 60,
                "unlock": 190, "clear": 120, "boosterclear": 120, "revive": 190
            }

        try:
            cfg = get_app_config(cur, app_id)
            if cfg and 'boosters' in cfg:
                for b in cfg['boosters']:
                    if isinstance(b, dict):
                        k = b.get('key','').replace('booster_','').replace('revive_','').lower()
                        if app_id != 2 and k in ['shuffle', 'undo', 'bubble', 'ufo']: continue
                        if app_id == 2 and k not in ['bubble', 'shuffle', 'ufo']: continue
                        BOOSTER_PRICES[k] = int(b.get('price', 100))
        except: pass

        # 2. QUERY
        where = "WHERE app_id = %s"; params = [app_id]
        if start_date: where += " AND created_at >= %s"; params.append(start_date + " 00:00:00")
        if end_date: where += " AND created_at <= %s"; params.append(end_date + " 23:59:59")
        
        cur.execute(f"SELECT created_at, event_name, event_json FROM event_logs {where} ORDER BY created_at ASC", tuple(params))
        all_rows = cur.fetchall()

        # 3. PROCESS
        metrics = {"start":0, "win":0, "fail":0, "spend":0, "rev":0}
        booster_counts = {}
        for k in BOOSTER_PRICES.keys(): booster_counts[k] = 0

        user_sessions = {} 
        win_costs = []
        fail_costs = []
        start_set = {"missionStart", "missionStart_Daily", "level_start", "level_loading_start"}
        win_set = {"missionComplete", "missionComplete_Daily", "level_win"}
        fail_set = {"missionFail", "missionFail_Daily", "level_fail", "level_lose"}
        EXCLUDED_KEYS = {'userID', 'uuid', 'user_id', 'session_id', 'event_timestamp', 'event_name', 'app_id', 'platform', 'os_ver', 'sdk_ver', 'device_id', 'levelID', 'level_display', 'missionID', 'dayChallenge', 'id', 'user_type', 'event_json', 'json', 'params', 'custom_attributes', 'value'}

        import re
        def clean_money(val):
            if not val: return 0
            s_val = str(val)
            if s_val.count('.') > 1: return 0
            if isinstance(val, (int, float)): return int(val)
            s = re.sub(r'[^\d]', '', str(val))
            return int(s) if s else 0

        target_lvl_int = int(level_id) if level_id and level_id.isdigit() else None
        processing_list = []

        # 4. FILTER
        for r in all_rows:
            json_str = r['event_json']
            data = universal_flatten(json_str)
            
            if target_lvl_int is not None:
                candidates = []
                for k in ['levelID', 'level_display', 'missionID', 'dayChallenge']:
                    val = data.get(k)
                    if val and str(val).isdigit(): candidates.append(int(val))
                if not candidates:
                    match = re.search(r'(?:levelID|level_display|missionID)[^0-9]{1,10}(\d+)', json_str)
                    if match: candidates.append(int(match.group(1)))
                if not candidates or max(candidates) != target_lvl_int: continue

            uid = data.get('userID') or data.get('uuid') or data.get('user_id') or "unknown"
            r['parsed'] = data
            r['temp_uid'] = uid 
            processing_list.append(r)

        # 5. SORT
        processing_list.sort(key=lambda x: (x['temp_uid'], x['created_at']))

        for r in processing_list:
            data = r['parsed']
            evt = r['event_name']
            uid = r['temp_uid']
            
            cost = 0
            raw_money = (data.get('coin_spent') or data.get('cost') or data.get('priceSpendLevel') or data.get('coinCost'))
            if not raw_money and evt.lower() == 'pricespendlevel':
                 for v in data.values():
                     v_c = clean_money(v)
                     if v_c > 0 and v_c < 100000: raw_money = v_c; break
            cost = clean_money(raw_money)

            if cost == 0:
                for k, v in data.items():
                    if ('booster' in k or 'revive' in k) and str(v).isdigit() and int(v) > 0:
                        clean_key = k.replace('booster_', '').replace('revive_', '').lower()
                        if clean_key in BOOSTER_PRICES:
                            cost += (int(v) * BOOSTER_PRICES[clean_key])

            if cost > 0:
                metrics['spend'] += 1
                metrics['rev'] += cost
                if uid in user_sessions: user_sessions[uid] += cost

            if evt in start_set:
                metrics['start'] += 1
                user_sessions[uid] = 0
            elif evt in win_set:
                metrics['win'] += 1
                if uid in user_sessions:
                    win_costs.append(user_sessions[uid])
                    del user_sessions[uid] 
            elif evt in fail_set:
                metrics['fail'] += 1
                if uid in user_sessions:
                    fail_costs.append(user_sessions[uid])
                    del user_sessions[uid]

            for k, v in data.items():
                if ('booster' in k or 'revive' in k) and str(v).isdigit() and int(v) > 0:
                    clean = k.replace('booster_', '').replace('revive_', '').lower()
                    if clean in BOOSTER_PRICES:
                        booster_counts[clean] += int(v)

        # 6. OUTPUT
        avg_win_cost = sum(win_costs) / len(win_costs) if win_costs else 0
        avg_fail_cost = sum(fail_costs) / len(fail_costs) if fail_costs else 0
        cost_arr = []
        if avg_win_cost > 0: cost_arr.append({"name": "Avg Win Cost", "value": int(avg_win_cost), "color": "#10b981"})
        if avg_fail_cost > 0: cost_arr.append({"name": "Avg Fail Cost", "value": int(avg_fail_cost), "color": "#ef4444"})
        if not cost_arr and metrics['rev'] > 0:
             cost_arr.append({"name": "Avg Spend per Play", "value": int(metrics['rev'] / (metrics['start'] or 1)), "color": "#f59e0b"})

        # [FIX V130] GOM NHÓM THEO TÊN HIỂN THỊ ĐỂ KHỬ TRÙNG LẶP
        merged_boosters = {}
        
        for k, cnt in booster_counts.items():
            pr = BOOSTER_PRICES.get(k, 100)
            display_name = DISPLAY_MAP.get(k, k.capitalize())
            
            # Khởi tạo nếu chưa có
            if display_name not in merged_boosters:
                merged_boosters[display_name] = {
                    "item_name": display_name, 
                    "usage_count": 0, 
                    "revenue": 0, 
                    "price": pr, 
                    "type": "Config"
                }
            
            # Cộng dồn chỉ số (Merge)
            merged_boosters[display_name]["usage_count"] += cnt
            merged_boosters[display_name]["revenue"] += (cnt * pr)
            
            # Nếu một trong các key thành phần có dùng -> Đánh dấu là Used
            if cnt > 0: merged_boosters[display_name]["type"] = "Used"

        # Chuyển về list và sort
        b_list = list(merged_boosters.values())
        b_list.sort(key=lambda x: x['usage_count'], reverse=True)

        processing_list.sort(key=lambda x: x['created_at'], reverse=True)
        total_rec = len(processing_list)
        paged_data = processing_list[offset : offset + limit]
        proc_logs = []
        
        for r in paged_data:
            d = r['parsed']
            details = []
            c_spent = d.get('coin_spent') or d.get('cost') or d.get('priceSpendLevel')
            if c_spent: details.append(f"💸 -{c_spent}")
            bal = d.get('coin_balance') or d.get('current_coin') or d.get('coinBalance') or d.get('gold_remain')
            if bal: details.append(f"💰 {bal}")
            
            for k, v in d.items():
                if k in EXCLUDED_KEYS or k in ['coin_spent', 'cost', 'priceSpendLevel', 'coin_balance', 'current_coin', 'gold_remain']: continue
                val_str = str(v)
                if val_str.startswith('{') or val_str == '0' or val_str == '': continue
                clean_k = k.replace('booster_', '').replace('revive_', '')
                details.append(f"{clean_k}: {val_str}")

            proc_logs.append({
                "time": r['created_at'].strftime('%H:%M:%S %d/%m'),
                "user_id": str(r['temp_uid'])[:8]+"..",
                "event_name": r['event_name'],
                "coin_spent": int(clean_money(c_spent) or 0),
                "item_name": " | ".join(details) if details else "-"
            })

        real_plays = metrics['win'] + metrics['fail']
        if real_plays == 0: real_plays = metrics['start']

        safe_response["metrics"] = {
            "total_plays": real_plays,
            "win_rate": round((metrics['win']/real_plays)*100, 1) if real_plays else 0,
            "arpu": metrics['rev'], 
            "avg_balance": 0, "top_item": b_list[0]['item_name'] if b_list else "None"
        }
        safe_response["funnel"] = [
            {"event_type":"START", "count":metrics['start'], "revenue":0},
            {"event_type":"WIN", "count":metrics['win'], "revenue":0},
            {"event_type":"FAIL", "count":metrics['fail'], "revenue":0}
        ]
        safe_response["booster_usage"] = b_list
        safe_response["cost_distribution"] = cost_arr
        safe_response["logs"] = {"data": proc_logs, "pagination": {"current": page, "total_pages": (total_rec+limit-1)//limit, "total_records": total_rec}}

        return jsonify(safe_response)

    except Exception as e:
        print(f"Level Detail Error V129: {e}")
        return jsonify(safe_response)
    finally: conn.close()

@app.route("/dashboard/<int:app_id>/strategic", methods=['GET'])
def get_strategic_overview(app_id):
    conn = get_db()
    if not conn: return jsonify({"success": False, "error": "DB error"}), 500

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        start_set = {"missionStart", "missionStart_Daily", "level_start", "level_loading_start", "missionStart_WeeklyQuestTutor"}
        fail_set = {"missionFail", "missionFail_Daily", "level_fail", "level_lose", "missionFail_WeeklyQuestTutor"}
        where = "WHERE app_id = %s"; params = [app_id]
        if start_date: where += " AND created_at >= %s"; params.append(start_date + " 00:00:00")
        if end_date: where += " AND created_at <= %s"; params.append(end_date + " 23:59:59")
        
        cur.execute(f"SELECT event_name, event_json FROM event_logs {where}", tuple(params))
        rows = cur.fetchall()
        
        stats = {} 

        for r in rows:
            data = universal_flatten(r['event_json'])
            
            # Tìm Level
            lvl_raw = data.get('levelID') or data.get('level_display') or data.get('missionID')
            if not lvl_raw: continue
            
            # Lấy số level
            digits = ''.join(filter(str.isdigit, str(lvl_raw)))
            if not digits: continue
            lvl_num = int(digits)
            if lvl_num > 2000: continue
            
            if lvl_num not in stats: stats[lvl_num] = {"plays": 0, "fails": 0, "rev": 0}
            
            evt = r['event_name']
            if evt in start_set: stats[lvl_num]['plays'] += 1
            elif evt in fail_set: stats[lvl_num]['fails'] += 1
            
            money = data.get('coin_spent') or data.get('coin_cost') or 0
            try: stats[lvl_num]['rev'] += float(money)
            except: pass

        chart = []
        for lvl, val in stats.items():
            # Chỉ hiện level có tương tác
            if val['plays'] > 0 or val['rev'] > 0 or val['fails'] > 0:
                fr = round((val['fails'] / val['plays']) * 100, 1) if val['plays'] > 0 else 0
                if fr > 100: fr = 100.0
                chart.append({
                    "name": f"Lv.{lvl}", "level_index": lvl,
                    "revenue": val['rev'], "fail_rate": fr, "plays": val['plays']
                })

        chart.sort(key=lambda x: x['level_index'])
        return jsonify({"success": True, "balance_chart": chart})

    except Exception as e:
        print(f"Strategic Error V97: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally: conn.close()

# --- BỔ SUNG API: XÓA 1 DÒNG & STOP JOB ---
@app.route("/monitor/history/<int:id>", methods=['DELETE'])
def delete_single_history(id):
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM job_history WHERE id = %s", (id,))
        conn.commit()
        return jsonify({"success": True, "msg": f"Deleted history #{id}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally: conn.close()

@app.route("/etl/stop/<int:hist_id>", methods=['POST'])
def stop_etl_process(hist_id):
    # API này dùng để đánh dấu job là Cancelled trên Database
    # Nó cũng cố gắng reset trạng thái bận của hệ thống nếu cần
    conn = get_db()
    try:
        # 1. Kích hoạt cờ dừng để Worker đang chạy tự thoát
        if hist_id in JOB_STOP_EVENTS:
            print(f"🛑 Sending STOP signal to Job #{hist_id}...")
            JOB_STOP_EVENTS[hist_id].set() # Đánh thức Worker ngay lập tức
        cur = conn.cursor()
        # Cập nhật DB
        cur.execute("""
            UPDATE job_history 
            SET status = 'Cancelled', end_time = NOW(), logs = logs || E'\n[USER MANUAL STOP]'
            WHERE id = %s AND status IN ('Running', 'Processing')
        """, (hist_id,))
        conn.commit()

        # 3. Reset hệ thống nếu cần
        if is_system_busy():
            set_system_busy(False)
            
        return jsonify({"success": True, "msg": f"Stop signal sent to Job #{hist_id}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally: conn.close()

# --- [V41 FIX] API CẤU HÌNH ĐỘNG (DÙNG JSON DB) ---
@app.route("/apps/<int:app_id>/analytics-config", methods=['GET', 'POST'])
def handle_analytics_config(app_id):
    conn = get_db()
    if not conn: return jsonify({"success": False, "error": "DB Connection Failed"}), 500
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # --- 1. LẤY CẤU HÌNH (GET) ---
        if request.method == 'GET':
            # Query mới: Chỉ lấy cột config_json
            cur.execute("SELECT config_json FROM analytics_config WHERE app_id = %s", (app_id,))
            row = cur.fetchone()
            
            if row and row['config_json']:
                # Trả về JSON chuẩn cho Frontend
                return jsonify(row['config_json'])
            else:
                # Nếu chưa có trong DB, trả về config mặc định từ hàm get_app_config
                # (Lưu ý: Bạn phải đảm bảo hàm get_app_config ở đầu file đã sửa tên bảng thành analytics_config nhé)
                return jsonify(get_app_config(cur, app_id))

        # --- 2. LƯU CẤU HÌNH (POST) ---
        elif request.method == 'POST':
            new_config = request.json # Frontend gửi lên toàn bộ cục JSON settings
            
            # Chuyển Dict thành String JSON để lưu vào DB
            config_str = json.dumps(new_config)

            # Lưu thẳng vào cột config_json (Gọn nhẹ hơn logic cũ rất nhiều)
            cur.execute("""
                INSERT INTO analytics_config (app_id, config_json, updated_at)
                VALUES (%s, %s, NOW())
                ON CONFLICT (app_id) DO UPDATE SET 
                    config_json = EXCLUDED.config_json,
                    updated_at = NOW()
            """, (app_id, config_str))
            
            conn.commit()
            return jsonify({"success": True, "msg": "Configuration Saved Successfully (V40 JSON Mode)"})
            
    except Exception as e:
        print(f"❌ Analytics Config Error: {e}")
        conn.rollback() # Chống kẹt transaction
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

# --- API CHẠY ETL (TỔNG HỢP DỮ LIỆU) ---
@app.route("/api/run-etl/<int:app_id>", methods=['POST'])
def trigger_etl_process(app_id):
    # Chạy trong thread riêng để không block server
    threading.Thread(target=run_etl_pipeline, args=(app_id,)).start()
    return jsonify({"status": "started", "message": "ETL process started in background"})

# --- API MỚI: TRA CỨU DỮ LIỆU THÔ (DATA EXPLORER) - FIXED TIMEZONE & PARAMS ---
@app.route("/events/search", methods=['GET'])
def search_events():
    try:
        app_id = request.args.get('app_id')
        try:
            page = int(request.args.get('page', 1))
            limit = int(request.args.get('limit', 20))
        except: page=1; limit=20
        
        # Các bộ lọc (ĐÃ THÊM PARSE_DATE_PARAM ĐỂ TRÁNH LỖI UNDEFINED)
        start_date = parse_date_param(request.args.get('start_date'))
        end_date = parse_date_param(request.args.get('end_date'))
        event_name = request.args.get('event_name')
        keyword = request.args.get('keyword') 
        level_filter = request.args.get('level')

        if not app_id:
            return jsonify({"success": False, "error": "Missing app_id"}), 400

        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. Xây dựng câu WHERE động
        where_clauses = ["app_id = %s"]
        params = [app_id]

        if start_date:
            # [FIX TIMEZONE] Ép DB convert sang giờ VN trước khi so sánh
            where_clauses.append("(created_at + interval '7 hours') >= %s")
            params.append(start_date + " 00:00:00")
        if end_date:
            # [FIX TIMEZONE] Ép DB convert sang giờ VN trước khi so sánh
            where_clauses.append("(created_at + interval '7 hours') <= %s")
            params.append(end_date + " 23:59:59")

        if event_name and event_name.strip():
            where_clauses.append("event_name = %s")
            params.append(event_name)

        if keyword and keyword.strip():
            where_clauses.append("event_json::text ILIKE %s")
            params.append(f"%{keyword}%")

        # Logic lọc Level (Regex)
        if level_filter and level_filter.strip():
            regex_pat = f'(levelID|level_display|missionID|dayChallenge)[^0-9a-zA-Z]+{level_filter}(\\D|$)'
            where_clauses.append("event_json::text ~* %s")
            params.append(regex_pat)

        full_where = " WHERE " + " AND ".join(where_clauses)

        # 2. Đếm tổng số dòng
        count_query = f"SELECT COUNT(*) as total FROM event_logs {full_where}"
        cursor.execute(count_query, tuple(params))
        total_records = cursor.fetchone()['total']
        total_pages = (total_records + limit - 1) // limit

        # 3. Lấy dữ liệu phân trang (CÓ FIX TIMEZONE)
        offset = (page - 1) * limit
        time_column = "to_char(created_at, 'DD/MM/YYYY HH24:MI:SS') || ' (VN: ' || to_char(created_at + interval '7 hours', 'DD/MM/YYYY HH24:MI:SS') || ')'"
        data_query = f"""
            SELECT id, event_name, {time_column} as created_at, event_json
            FROM event_logs 
            {full_where}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """
        
        # [AN TOÀN]: Tạo list mới cho query này thay vì extend list cũ để tránh lỗi "Sốc thuốc"
        final_params = params + [limit, offset]
        
        cursor.execute(data_query, tuple(final_params))
        rows = cursor.fetchall()

        # 4. Xử lý hiển thị (Key Info)
        import json
        for row in rows:
            try:
                # Parse JSON đa năng
                raw = row['event_json']
                parsed = {}

                if isinstance(raw, str):
                    try: parsed = json.loads(raw)
                    except: parsed = {}
                elif isinstance(raw, dict):
                    parsed = raw 
                
                # Flatten double-encoded json
                if isinstance(parsed, dict) and 'event_json' in parsed and isinstance(parsed['event_json'], str):
                    try:
                        inner = json.loads(parsed['event_json'])
                        parsed.update(inner)
                    except: pass
                
                row['event_json'] = parsed
                data = parsed 
                
                # --- TẠO KEY INFO (CONTEXT) ---
                info_parts = []
                
                # A. Level & Mission
                lvl = data.get('levelID') or data.get('level_display') or data.get('dayChallenge')
                if lvl: info_parts.append(f"Lv.{lvl}")
                
                mis = data.get('missionID')
                if mis and str(mis) != str(lvl): 
                    info_parts.append(f"Ms.{mis}")
                
                # B. Tiền tệ
                coin = 0
                if 'coin_spent' in data: coin = -int(data['coin_spent'])
                elif 'coin_cost' in data: coin = -int(data['coin_cost'])
                elif 'priceSpendLevel' in data: coin = -int(data['priceSpendLevel'])
                elif 'coinBalance' in data: coin = f"Bal:{data['coinBalance']}"
                elif 'gold_remain' in data: coin = f"Gold:{data['gold_remain']}"
                elif 'goldRemain' in data:  coin = f"Gold:{data['goldRemain']}"
                
                if isinstance(coin, int) and coin != 0: info_parts.append(f"{coin} Coin")
                elif isinstance(coin, str): info_parts.append(coin)

                # C. Mạng sống (Game 2)
                if 'life_remain' in data:
                    info_parts.append(f"❤️ {data['life_remain']}")

                if 'revenue' in data: info_parts.append(f"+${data['revenue']}")

                # D. Boosters & Items (Game 2)
                boosters_used = []
                for k, v in data.items():
                    if k.startswith('booster_') and str(v).isdigit() and int(v) > 0:
                        clean_name = k.replace('booster_', '').capitalize()
                        if clean_name == 'Bubble': clean_name = '🫧 Bubble'
                        if clean_name == 'Shuffle': clean_name = '🔀 Shuffle'
                        if clean_name == 'Ufo': clean_name = '🛸 UFO'
                        boosters_used.append(f"{clean_name} x{v}")
                
                if boosters_used:
                    info_parts.append(" | ".join(boosters_used))

                # E. Item Name
                if 'item_name' in data: info_parts.append(f"🛒 {data['item_name']}")

                row['key_info'] = " | ".join(info_parts) if info_parts else "..."
            
            except Exception as e:
                row['key_info'] = '-'

        return jsonify({
            "success": True,
            "data": rows,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "total_records": total_records,
                "limit": limit
            }
        })

    except Exception as e:
        print(f"Search Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

# --- API DATA CHECK (PHIÊN BẢN HỒI SỨC: DEEP UNPACK + PYTHON FILTER) ---
@app.route("/api/data-check/<int:app_id>", methods=['GET'])
def get_data_check(app_id):
    conn = get_db()
    if not conn: return jsonify({"success": False}), 500
    
    # 1. Lấy tham số Filter
    start_date = parse_date_param(request.args.get('start_date'))
    end_date = parse_date_param(request.args.get('end_date'))
    filter_ver = request.args.get('version')
    filter_geo = request.args.get('geo')

    # 2. Cấu hình Giá (Game 2)
    PRICES = {
        "bubble": 800, "shuffle": 800, "ufo": 1000, "balloon": 800, "changeHole": 1000,
        "revive": 700, "unlock": 190
    }

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # 3. Query DB (Chỉ lọc theo Ngày + App ID để tối ưu index)
        where_clauses = ["app_id = %s"]
        params = [app_id]

        if start_date:
            # [FIX TIMEZONE] Ép DB convert sang giờ VN trước khi so sánh
            where_clauses.append("(created_at + interval '7 hours') >= %s")
            params.append(start_date + " 00:00:00")
        if end_date:
            # [FIX TIMEZONE] Ép DB convert sang giờ VN trước khi so sánh
            where_clauses.append("(created_at + interval '7 hours') <= %s")
            params.append(end_date + " 23:59:59")
        
        full_where = " AND ".join(where_clauses)
        
        cur.execute(f"""
            SELECT event_name, event_json 
            FROM event_logs 
            WHERE {full_where}
        """, tuple(params))
        
        rows = cur.fetchall()

        # 4. Xử lý dữ liệu (Aggregation & Deep Parse)
        stats = {}
        all_boosters_found = set() # CHÌA KHÓA MỚI: Rổ hứng mọi loại Booster trên đời
        start_events = {"level_start", "missionStart", "missionStart_Daily", "level_first_start"}
        win_events = {"level_win", "missionComplete", "missionComplete_Daily", "level_first_end"}
        end_events = win_events.union({"level_lose", "missionFail", "missionFail_Daily"})
        
        import json

        for r in rows:
            evt_name = r['event_name']
            raw_json = r['event_json']
            
            data = {}
            if isinstance(raw_json, str):
                try: data = json.loads(raw_json)
                except: continue
            else: data = raw_json if raw_json else {}

            if 'event_json' in data and isinstance(data['event_json'], str):
                try:
                    inner = json.loads(data['event_json'])
                    if isinstance(inner, dict): data.update(inner) 
                except: pass

            if filter_ver and filter_ver != 'all' and str(data.get('app_version_name', '')) != str(filter_ver): continue
            if filter_geo and filter_geo != 'all' and str(data.get('country_iso_code', '')) != str(filter_geo): continue

            lvl = None
            for k in ['level_display', 'levelID', 'missionID', 'level']:
                if k in data and str(data[k]).isdigit():
                    lvl = int(data[k])
                    break
            
            if lvl is None: continue 
            
            # Khởi tạo struct (Booster rỗng để hứng tự động)
            if lvl not in stats:
                stats[lvl] = {
                    "user_start_set": set(), "user_win_set": set(), "total_plays": 0,
                    "boosters": {}, # Không hardcode bubble/shuffle nữa
                    "timeplay_sum": 0, "timeplay_count": 0, "total_revive": 0
                }
            
            s = stats[lvl]
            uid = data.get('uuid') or data.get('userID') or data.get('device_id') or "unknown"

            if evt_name in start_events:
                s['user_start_set'].add(uid) 
                if evt_name != "level_first_start": s['total_plays'] += 1
            
            elif evt_name in win_events:
                s['user_win_set'].add(uid)
                if evt_name != "level_first_end":
                    tm = data.get('timeplay') or data.get('duration') or data.get('timePlay')
                    if tm:
                        try: 
                            val = float(tm)
                            if 0 < val < 7200: 
                                s['timeplay_sum'] += val
                                s['timeplay_count'] += 1
                        except: pass

            if evt_name in end_events:
                # Dò Tự Động Mọi Loại Booster + BỘ LỌC RÁC
                IGNORED_BOOSTERS = ['carpaint'] # Danh sách đen các booster không dùng nữa
                BOOSTER_MAP = {
                'ufo': 'changehole',
                'shuffle': 'balloon'
                }
            # DÒ TỰ ĐỘNG + BỘ LỌC RÁC + ĐỒNG BỘ TÊN (ALIAS MAPPING)
            IGNORED_BOOSTERS = ['carpaint'] # Chỉ vứt rác thật sự
            
            # SỔ THÔNG DỊCH: Map tên cũ (key) sang tên mới (value)
            BOOSTER_MAP = {
                'ufo': 'changehole',
                'shuffle': 'balloon'
            }

            for k, v in data.items():
                if k.startswith('booster_') and str(v).lstrip('-').isdigit():
                    clean_k = k.replace('booster_', '').lower()
                    
                    # 1. Quét radar xem có phải rác không
                    if clean_k in IGNORED_BOOSTERS:
                        continue
                        
                    # 2. [MA THUẬT Ở ĐÂY]: Phiên dịch tên cũ sang tên mới
                    if clean_k in BOOSTER_MAP:
                        clean_k = BOOSTER_MAP[clean_k]
                        
                    # 3. Bỏ vào rổ và đếm số
                    all_boosters_found.add(clean_k)
                    if int(v) > 0:
                        s['boosters'][clean_k] = s['boosters'].get(clean_k, 0) + int(v)

        # 5. Tổng hợp báo cáo
        report = []
        sorted_levels = sorted(stats.keys())
        
        for i, lvl in enumerate(sorted_levels):
            s = stats[lvl]
            u_start = len(s['user_start_set'])
            u_win = len(s['user_win_set'])
            if u_start < u_win: u_start = u_win 
            
            level_drop = round(((u_start - u_win) / u_start) * 100, 2) if u_start > 0 else 0
            play_count = round(s['total_plays'] / u_start, 2) if u_start > 0 else 0

            next_drop = 0 
            if i > 0:
                prev_lvl = sorted_levels[i-1]
                win_prev_set = stats[prev_lvl]['user_win_set']
                start_current_set = s['user_start_set']
                if len(win_prev_set) > 0:
                    drop_users = win_prev_set - start_current_set
                    next_drop = round((len(drop_users) / len(win_prev_set)) * 100, 2)

            total_b = sum(s.get('boosters', {}).values())
            total_r = s.get('total_revive', 0)
            
            # Tính tiền tự động cho cả các loại booster mới
            coin_spent = total_r * PRICES.get('revive', 700)
            for b_name, b_qty in s.get('boosters', {}).items():
                coin_spent += b_qty * PRICES.get(b_name.lower(), 100) # Nếu không có trong PRICES, mặc định là 100
            
            avg_b = round((total_b + total_r) / u_start, 2) if u_start > 0 else 0
            avg_c = round(coin_spent / u_start, 0) if u_start > 0 else 0
            avg_t = round(s.get('timeplay_sum', 0) / s.get('timeplay_count', 1), 1) if s.get('timeplay_count', 0) > 0 else 0

            row_dict = {
                "level": lvl,
                "user_start": u_start,
                "user_win": u_win,
                "level_drop": level_drop,
                "next_drop": next_drop,
                "play_count": play_count,
                "total_booster": total_b,
                "total_revive": total_r,
                "avg_booster": avg_b,
                "avg_coin": avg_c,
                "avg_time": avg_t,
                "boosters": s.get('boosters', {})
            }
            
            # [BÁC SĨ FIX]: Đổ đồng loạt số 0 ra Frontend nếu không dùng
            for b in all_boosters_found:
                row_dict[f"booster_{b.lower()}"] = s.get('boosters', {}).get(b, 0)

            report.append(row_dict)

        return jsonify({"success": True, "data": report})

    except Exception as e:
        print(f"DataCheck Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally: conn.close()

# --- API TRUY TÌM NGƯỜI CHƠI BỎ (DROPPED USERS) ---
@app.route("/api/dropped-users/<int:app_id>", methods=['GET'])
def get_dropped_users(app_id):
    conn = get_db()
    if not conn: return jsonify({"success": False}), 500

    target_level = request.args.get('level', type=int)
    start_date = parse_date_param(request.args.get('start_date'))
    end_date = parse_date_param(request.args.get('end_date'))
    filter_ver = request.args.get('version')

    if target_level is None:
        return jsonify({"success": False, "error": "Chưa nhập level cần soi"}), 400

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        where_clauses = ["app_id = %s"]
        params = [app_id]

        if start_date:
            where_clauses.append("(created_at + interval '7 hours') >= %s")
            params.append(start_date + " 00:00:00")
        if end_date:
            where_clauses.append("(created_at + interval '7 hours') <= %s")
            params.append(end_date + " 23:59:59")

        full_where = " AND ".join(where_clauses)
        
        # Tối ưu: Chỉ lôi nhóm Start và Win lên RAM để tính toán cho nhanh
        cur.execute(f"""
            SELECT event_name, event_json 
            FROM event_logs 
            WHERE {full_where}
              AND event_name IN ('level_start', 'missionStart', 'missionStart_Daily', 'level_first_start', 'level_win', 'missionComplete', 'missionComplete_Daily', 'level_first_end')
        """, tuple(params))
        
        rows = cur.fetchall()
        start_set = set()
        win_set = set()
        
        import json
        for r in rows:
            evt_name = r['event_name']
            raw_json = r['event_json']
            
            data = {}
            if isinstance(raw_json, str):
                try: data = json.loads(raw_json)
                except: continue
            else: data = raw_json if raw_json else {}

            # Giải nén sâu lấy UID thật (Mũi tiêm làm nên sự khác biệt)
            if 'event_json' in data and isinstance(data['event_json'], str):
                try:
                    inner = json.loads(data['event_json'])
                    if isinstance(inner, dict): data.update(inner)
                except: pass

            if filter_ver and filter_ver != 'all':
                if str(data.get('app_version_name', '')) != str(filter_ver): continue

            # Xác định Level
            lvl = None
            for k in ['level_display', 'levelID', 'missionID', 'level']:
                if k in data and str(data[k]).isdigit():
                    lvl = int(data[k])
                    break
            
            if lvl != target_level: continue 
            
            # Khóa mục tiêu UID
            uid = data.get('uuid') or data.get('userID') or data.get('device_id') or "unknown"
            
            if evt_name in {"level_start", "missionStart", "missionStart_Daily", "level_first_start"}:
                start_set.add(uid)
            elif evt_name in {"level_win", "missionComplete", "missionComplete_Daily", "level_first_end"}:
                win_set.add(uid)

        # PHÉP THUẬT Ở ĐÂY: Lấy tập Start trừ tập Win
        dropped_uuids = list(start_set - win_set)
        
        return jsonify({
            "success": True, 
            "level": target_level,
            "total_start": len(start_set),
            "total_win": len(win_set),
            "dropped_count": len(dropped_uuids),
            "dropped_uuids": dropped_uuids # Trả thẳng danh sách mã UID ra ngoài
        })

    except Exception as e: return jsonify({"success": False, "error": str(e)}), 500
    finally: conn.close()

# API GET EVENT DICTIONARY (SETTINGS TAB)
@app.route("/api/events/dictionary/<int:app_id>", methods=['GET'])
def get_event_dictionary(app_id):
    conn = get_db()
    if not conn: return jsonify({})
    
    try:
        cur = conn.cursor()
        # Chỉ lấy tên event duy nhất
        cur.execute("SELECT DISTINCT event_name FROM event_logs WHERE app_id = %s ORDER BY event_name", (app_id,))
        rows = cur.fetchall()
        
        events = [r[0] for r in rows]
        
        # Logic phân loại thủ công
        groups = {
            "Progression 🎯": [], # Level, Mission...
            "Economy & IAP 💎": [], # Coin, Gem, Shop, Purchase...
            "System & Tech ⚙️": [], # Login, Loading, Error...
            "Ads & Rewards 🎬": [], # Ad, Reward...
            "Others 📦": []         # Còn lại
        }

        for e in events:
            low = e.lower()
            if any(x in low for x in ['level', 'mission', 'quest', 'stage', 'checkpoint', 'tutorial', 'win', 'fail', 'lose', 'complete', 'start']):
                groups["Progression 🎯"].append(e)
            elif any(x in low for x in ['iap', 'purchase', 'coin', 'gold', 'gem', 'money', 'shop', 'buy', 'spend', 'cost', 'price']):
                groups["Economy & IAP 💎"].append(e)
            elif any(x in low for x in ['ad_', 'ads', 'reward', 'bonus']):
                groups["Ads & Rewards 🎬"].append(e)
            elif any(x in low for x in ['login', 'session', 'load', 'init', 'install', 'update', 'error', 'ping']):
                groups["System & Tech ⚙️"].append(e)
            else:
                groups["Others 📦"].append(e)

        # Xóa các nhóm rỗng để đỡ rối
        final_groups = {k: v for k, v in groups.items() if v}
        
        return jsonify({
            "success": True, 
            "total_count": len(events),
            "groups": final_groups
        })

    except Exception as e:
        print(f"Error Event Dict: {e}")
        return jsonify({})
    finally: conn.close()

# --- API EXPORT EXCEL (NÂNG CẤP V2: ĐỒNG BỘ LOGIC DATA CHECK) ---
@app.route("/api/datacheck/export/<int:app_id>", methods=['GET'])
def export_datacheck_excel(app_id):
    import io
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter 

    conn = get_db()
    if not conn: return jsonify({"success": False}), 500
    
    # 1. Lấy Filter (Đồng bộ với FE)
    start_date = parse_date_param(request.args.get('start_date'))
    end_date = parse_date_param(request.args.get('end_date'))
    filter_ver = request.args.get('version')
    filter_geo = request.args.get('geo')
    
    # 2. Config Giá
    PRICES = {
        "bubble": 800, "shuffle": 800, "ufo": 1000, "balloon": 800, "changeHole": 1000,
        "revive": 700, "unlock": 190
    }
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # 3. Query DB
        where_clauses = ["app_id = %s"]
        params = [app_id]
        if start_date:
            # [FIX TIMEZONE] Ép DB convert sang giờ VN trước khi so sánh
            where_clauses.append("(created_at + interval '7 hours') >= %s")
            params.append(start_date + " 00:00:00")
        if end_date:
            # [FIX TIMEZONE] Ép DB convert sang giờ VN trước khi so sánh
            where_clauses.append("(created_at + interval '7 hours') <= %s")
            params.append(end_date + " 23:59:59")
            
        full_where = " AND ".join(where_clauses)
        cur.execute(f"SELECT event_name, event_json FROM event_logs WHERE {full_where}", tuple(params))
        rows = cur.fetchall()

        # 4. Xử lý dữ liệu (Aggregation & Deep Parse)
        stats = {}
        all_boosters_found = set()
        start_events = {"level_start", "missionStart", "missionStart_Daily", "level_first_start"}
        win_events = {"level_win", "missionComplete", "missionComplete_Daily", "level_first_end"}

        for r in rows:
            evt_name = r['event_name']
            raw_json = r['event_json']
            
            data = {}
            if isinstance(raw_json, str):
                try: data = json.loads(raw_json)
                except: continue
            else: data = raw_json if raw_json else {}

            if 'event_json' in data and isinstance(data['event_json'], str):
                try:
                    inner = json.loads(data['event_json'])
                    if isinstance(inner, dict): data.update(inner)
                except: pass

            if filter_ver and filter_ver != 'all' and str(data.get('app_version_name', '')) != str(filter_ver): continue
            if filter_geo and filter_geo != 'all' and str(data.get('country_iso_code', '')) != str(filter_geo): continue

            lvl = None
            for k in ['level_display', 'levelID', 'missionID', 'level']:
                if k in data and str(data[k]).isdigit():
                    lvl = int(data[k])
                    break
            if lvl is None or lvl > 10000: continue 

            if lvl not in stats:
                stats[lvl] = {
                    "user_start_set": set(), "user_win_set": set(), "total_plays": 0,
                    "boosters": {}, 
                    "total_revive": 0,
                    "timeplay_sum": 0, "timeplay_count": 0
                }
            s = stats[lvl]
            uid = data.get('uuid') or data.get('userID') or "unknown"

            if evt_name in start_events:
                s['user_start_set'].add(uid)
                if evt_name != "level_first_start": s['total_plays'] += 1
            elif evt_name in win_events:
                s['user_win_set'].add(uid)
                if evt_name != "level_first_end":
                    tm = data.get('timeplay') or data.get('duration')
                    if tm:
                        try: 
                            val = float(tm)
                            if 0 < val < 7200: 
                                s['timeplay_sum'] += val
                                s['timeplay_count'] += 1
                        except: pass

            # Đếm Booster Động
            IGNORED_BOOSTERS = ['carpaint'] # BỘ LỌC RÁC: Thêm các booster muốn loại bỏ vào đây
            BOOSTER_MAP = {
                'ufo': 'changehole',
                'shuffle': 'balloon'
            }
            for k, v in data.items():
                if k.startswith('booster_') and str(v).lstrip('-').isdigit():
                    clean_k = k.replace('booster_', '').lower()
                    
                    # 1. Quét radar xem có phải rác không
                    if clean_k in IGNORED_BOOSTERS:
                        continue
                        
                    # 2. [MA THUẬT Ở ĐÂY]: Phiên dịch tên cũ sang tên mới
                    if clean_k in BOOSTER_MAP:
                        clean_k = BOOSTER_MAP[clean_k]
                        
                    # 3. Bỏ vào rổ và đếm số
                    all_boosters_found.add(clean_k)
                    if int(v) > 0:
                        s['boosters'][clean_k] = s['boosters'].get(clean_k, 0) + int(v)

        # 5. TẠO FILE EXCEL
        wb = Workbook(); ws = wb.active; ws.title = "Data Check Report"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        RED_FONT = Font(color="DC2626", bold=True)
        ORANGE_FONT = Font(color="D97706", bold=True)
        
        # Cột Booster lấy tự động từ rổ
        booster_keys = sorted(list(all_boosters_found))

        headers = [
            "Level", "User Start", "User Win", "Level Drop (%)", "Next Drop (%)", 
            "Drop Change", "Play Count", "Unlock"
        ]
        for k in booster_keys:
            headers.append(f"Booster {k.capitalize()}")
            
        headers.extend([
            "Total Booster Used", 
            "Revive Full", "Revive Moves", "Total Revive", 
            "Avg B&R / User", "Avg Time (s)", "Coin Spend Diff", "Avg Coin Spend"
        ])
        
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border

        sorted_levels = sorted(stats.keys())
        for i, lvl in enumerate(sorted_levels):
            s = stats[lvl]
            u_start = len(s['user_start_set'])
            u_win = len(s['user_win_set'])
            if u_start < u_win: u_start = u_win
            
            level_drop = round(((u_start - u_win) / u_start) * 100, 2) if u_start > 0 else 0
            play_count = round(s['total_plays'] / u_start, 2) if u_start > 0 else 0
            
            next_drop = 0
            if i > 0:
                prev_lvl = sorted_levels[i-1]
                win_prev_set = stats[prev_lvl]['user_win_set']
                start_current_set = s['user_start_set']
                if len(win_prev_set) > 0:
                    drop_users = win_prev_set - start_current_set
                    next_drop = round((len(drop_users) / len(win_prev_set)) * 100, 2)

            total_b = sum(s['boosters'].values())
            total_r = s['total_revive']
            
            coin_spent = total_r * PRICES.get('revive', 700)
            for b_name, b_qty in s['boosters'].items():
                coin_spent += b_qty * PRICES.get(b_name.lower(), 100)
            
            avg_b = round((total_b + total_r) / u_start, 2) if u_start > 0 else 0
            avg_c = round(coin_spent / u_start, 0) if u_start > 0 else 0
            avg_t = round(s['timeplay_sum'] / s['timeplay_count'], 1) if s['timeplay_count'] > 0 else 0

            # [BÁC SĨ FIX]: Đổ số 0 thẳng thừng thay vì dấu "-"
            row_data = [
                lvl, u_start, u_win, f"{level_drop}%", 
                f"{next_drop}%" if next_drop > 0 else 0,
                0, play_count, 0  
            ]
            
            for k in booster_keys:
                val = s['boosters'].get(k, 0)
                row_data.append(val) # Hiện số 0 rành rành
                
            row_data.extend([
                total_b,
                0, 0, total_r, 
                avg_b, avg_t, 0, avg_c
            ])
            
            ws.append(row_data)
            
            curr_row = ws.max_row
            cell_drop = ws.cell(row=curr_row, column=4)
            if level_drop > 50: cell_drop.font = RED_FONT
            elif level_drop > 30: cell_drop.font = ORANGE_FONT
            for cell in ws[curr_row]:
                cell.alignment = center_align
                cell.border = thin_border

        for i, col in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"DataCheck_App{app_id}_{start_date}_{end_date}.xlsx"
        from flask import send_file
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

    except Exception as e:
        import traceback
        print(f"Export V2 Error: {traceback.format_exc()}")
        return f"Error: {e}", 500
    finally: conn.close()

@app.route("/monitor/export/<int:job_id>", methods=['GET'])
def export_job_raw_data(job_id):
    import io
    import json
    from datetime import timedelta
    from flask import send_file # Đảm bảo đã import cái này
    
    conn = get_db()
    if not conn: return jsonify({"error": "DB Disconnected"}), 500
    
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. Lấy thông tin Job
        cur.execute("SELECT app_id, date_since, date_until FROM job_history WHERE id = %s", (job_id,))
        job = cur.fetchone()
        
        if not job: return "Job not found", 404
        if not job['date_since'] or not job['date_until']: return "Job has no time range data", 400

        # 2. CHIẾN THUẬT QUÉT KÉP (DUAL SCAN STRATEGY)
        raw_start = job['date_since']
        raw_end = job['date_until']
        
        candidates = [
            (raw_start, raw_end, "Original (UTC stored)"),
            (raw_start - timedelta(hours=7), raw_end - timedelta(hours=7), "Corrected (VN stored)")
        ]
        
        rows = []
        used_strategy = "None"

        for start, end, label in candidates:
            # === THAY ĐỔI QUAN TRỌNG: SELECT * ĐỂ LẤY HẾT CỘT ===
            # Thay vì chỉ chọn vài cột, ta lấy tất cả (SELECT *)
            cur.execute("""
                SELECT * FROM event_logs 
                WHERE app_id = %s 
                  AND created_at >= %s 
                  AND created_at <= %s
                ORDER BY created_at ASC
            """, (job['app_id'], start, end))
            
            result = cur.fetchall()
            if len(result) > 0:
                rows = result
                used_strategy = label
                print(f"✅ Export Job #{job_id}: Found {len(rows)} events using strategy: {label}")
                break 
        
        # 3. Parse & Đóng gói (Logic mới: Trộn tất cả vào 'data')
        export_data = []
        for r in rows:
            # A. Xử lý phần lõi JSON (Game logic - level, gold...)
            try:
                core_data = json.loads(r['event_json']) if isinstance(r['event_json'], str) else r['event_json']
                if not isinstance(core_data, dict): core_data = {} 
            except:
                core_data = {}
            
            # B. Xử lý phần vỏ (Metadata: device, city, os, mcc, mnc...)
            # Chúng ta sẽ nhét tất cả các cột của DB vào trong biến 'data' luôn
            merged_data = core_data.copy() # Bắt đầu với dữ liệu game
            
            for key, val in r.items():
                # Loại bỏ các trường hệ thống không cần thiết hoặc trùng lặp
                if key in ['id', 'event_json', 'created_at', 'updated_at', 'app_id', 'job_id']:
                    continue
                
                # Format lại các trường ngày tháng nếu có (tránh lỗi JSON serializable)
                if hasattr(val, 'isoformat'):
                    val = val.isoformat()
                    
                # Ghi đè hoặc thêm mới vào merged_data
                # Đây là bước quan trọng: Nó đưa city, mcc, mnc... vào chung với dữ liệu game
                merged_data[key] = val

            # C. Tạo cấu trúc bản ghi cuối cùng
            # Lấy timestamp đẹp để hiển thị bên ngoài (dùng created_at của DB làm chuẩn)
            ts_str = r['created_at'].strftime('%Y-%m-%d %H:%M:%S') if r.get('created_at') else ""

            export_data.append({
                "time": ts_str,
                "name": r.get('event_name', 'unknown'),
                "data": merged_data # <--- ĐÂY LÀ CỤC DỮ LIỆU ĐẦY ĐỦ NHẤT (FULL OPTION)
            })

        # 4. Trả file
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
        mem_file = io.BytesIO(json_str.encode('utf-8'))
        
        filename = f"Job_{job_id}_RawData_Full_{used_strategy.split()[0]}.json"
        
        return send_file(
            mem_file, 
            mimetype='application/json',
            as_attachment=True, 
            download_name=filename
        )

    except Exception as e:
        print(f"Export Job Error: {e}")
        return f"Error: {e}", 500
    finally:
        if conn: conn.close()

# --- [API TỐI ƯU] LẤY FILTER VERSION & GEO BẰNG SQL (FIX LỖI WARNING) ---
@app.route("/api/filters/options/<int:app_id>", methods=['GET'])
def get_filter_options(app_id):
    conn = get_db()
    if not conn: return jsonify({"versions": ["All"], "geos": ["All"]})
    
    try:
        cur = conn.cursor()
        
        # --- CÁCH 1: LẤY VERSION ---
        # [FIX]: Thêm chữ r trước dấu ngoặc kép ba (r""") để Python không báo lỗi \s
        cur.execute(r"""
            SELECT DISTINCT substring(event_json FROM '"app_version_name":\s*"([^"]+)"') as ver
            FROM event_logs 
            WHERE app_id = %s 
              AND event_json LIKE '%%app_version_name%%'
        """, (app_id,))
        
        raw_versions = cur.fetchall()
        versions = sorted([r[0] for r in raw_versions if r[0]], reverse=True)

        # --- CÁCH 2: LẤY GEO ---
        # [FIX]: Thêm chữ r ở đây nữa
        cur.execute(r"""
            SELECT DISTINCT substring(event_json FROM '"country_iso_code":\s*"([^"]+)"') as geo
            FROM event_logs 
            WHERE app_id = %s 
              AND event_json LIKE '%%country_iso_code%%'
        """, (app_id,))
        
        raw_geos = cur.fetchall()
        geos = sorted([r[0] for r in raw_geos if r[0]])
        
        return jsonify({
            "versions": ["All"] + versions,
            "geos": ["All"] + geos
        })
        
    except Exception as e:
        print(f"Filter Option Error (SQL Mode): {e}")
        return jsonify({"versions": ["All"], "geos": ["All"]})
    finally:
        conn.close()

@app.route("/api/create_manual_job", methods=['POST'])
def create_manual_job():
    conn = get_db()
    if not conn: return jsonify({"error": "DB Disconnected"}), 500
    
    try:
        data = request.json
        app_id = data.get('app_id')
        start_time_str = data.get('start_time') 
        end_time_str = data.get('end_time')
        execution_time_str = data.get('execution_time') 
        
        if not app_id or not start_time_str or not end_time_str:
            return jsonify({"error": "Thiếu thông tin bắt buộc"}), 400

        # [FIX TIMEZONE]: Input từ Frontend là giờ VN -> Trừ 7 tiếng để lưu UTC vào DB
        # AppMetrica cần giờ UTC, và DB chúng ta quy ước lưu UTC
        dt_start = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M') - timedelta(hours=7)
        dt_end = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M') - timedelta(hours=7)
        
        dt_scheduled = None
        if execution_time_str:
             # Giờ hẹn chạy cũng phải trừ 7 tiếng để Worker (chạy giờ server UTC/System) hiểu đúng
             dt_scheduled = datetime.strptime(execution_time_str, '%Y-%m-%dT%H:%M') - timedelta(hours=7)

        if dt_end <= dt_start:
            return jsonify({"error": "Data End Time phải lớn hơn Start Time!"}), 400

        # Chèn vào DB với run_type='manual'
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO job_history 
            (app_id, date_since, date_until, status, retry_count, created_at, scheduled_at, logs, run_type)
            VALUES (%s, %s, %s, 'pending', 0, NOW(), %s, 'Manual Job Scheduled by User', 'manual')
            RETURNING id;
        """, (app_id, dt_start, dt_end, dt_scheduled))
        
        new_job_id = cur.fetchone()[0]
        conn.commit()
        
        msg = "Đã lên lịch (Scheduled)!" if dt_scheduled else "Đã tạo Job, chạy ngay!"
        return jsonify({"message": msg, "job_id": new_job_id})

    except Exception as e:
        conn.rollback()
        print(f"Manual Job Error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

if __name__ == '__main__':
    # 1. Kích hoạt Cô Y Tá (Scheduler)
    t_scheduler = threading.Thread(target=run_scheduler_loop)
    t_scheduler.daemon = True
    t_scheduler.start()

    # 2. [PHÉP THUẬT ĐA LUỒNG]: Tuyển thẳng 3 Bác Sĩ chạy song song
    NUM_WORKERS = 3
    worker_threads = []
    
    for i in range(NUM_WORKERS):
        t_worker = threading.Thread(target=run_worker_loop, args=(i+1,))
        t_worker.daemon = True
        t_worker.start()
        worker_threads.append(t_worker)

    print(f"🚀 SYSTEM READY: Smart Scheduler & {NUM_WORKERS} Worker Threads started...")
    
    backend_port = int(os.getenv("BACKEND_PORT", "8080"))
    print(f"🌐 Backend running on 127.0.0.1:{backend_port}")
    app.run(port=backend_port, debug=True, use_reloader=False)